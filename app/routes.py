# ==========================
# IMPORTS PADRÃO PYTHON
# ==========================
import os
import re
import tempfile
import unicodedata
from datetime import date, datetime
from io import BytesIO
import json


from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import uuid
import os

# ==========================
# FLASK
# ==========================
from flask import (Blueprint, after_this_request, current_app, flash, jsonify,
                   redirect, render_template, request, send_file,
                   send_from_directory, url_for)

from flask_login import current_user , login_required

# ==========================
# EXCEL / PDF
# ==========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape
# ==========================
# SQLALCHEMY / BANCO
# ==========================
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

# ==========================
# APP
# ==========================
from app import db
from app.models import Notificacao, Solicitacao, Usuario, Clientes, Pilotos

print("--- ROTAS CARREGADAS COM SUCESSO ---")

bp = Blueprint('main', __name__)

# --- 1: GLOBAL CONTEXT  ---
@bp.context_processor
def inject_globals():
    """
    Otimizado para não travar a navegação. 
    Faz apenas uma consulta de contagem simples por página carregada.
    """
    if current_user.is_authenticated:
        # Consulta de contagem direta no banco (muito mais rápida que carregar objetos)
        q = db.session.query(db.func.count(Notificacao.id)).filter(
            Notificacao.lida_em.is_(None),
            Notificacao.apagada_em.is_(None)
        )
        
        if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
            q = q.filter(Notificacao.usuario_id == current_user.id)
            
        return dict(notif_count=q.scalar() or 0)
    return dict(notif_count=0)


# --- 2: FILTRO DE DATA PARA JINJA2 ---
@bp.app_template_filter('datetimeformat')
def datetimeformat(value, format='%d-%m-%y'):
    """
    Filtro para formatar datas no Jinja2.
    Otimizado para lidar com strings, objetos datetime e valores nulos.
    """
    if value is None:
        return ""
    try:
        if isinstance(value, str):
            # Se for string (ex: '2025-12-31'), converte para objeto date antes de formatar
            return datetime.strptime(value, "%Y-%m-%d").strftime(format)
        return value.strftime(format)
    except Exception:
        return value # Retorna o valor original em caso de erro para não quebrar a página

def get_upload_folder():
    """
    Localiza a pasta de uploads de forma absoluta.
    Garante que a pasta exista sem processamento repetitivo desnecessário.
    """
    # Pasta 'upload-files' no mesmo nível da pasta 'app'
    folder = os.path.join(current_app.root_path, '..', 'upload-files')
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    return os.path.abspath(folder)

import unicodedata

def normalize_string(value):
    if value:
        return ''.join(
            c for c in unicodedata.normalize('NFD', value)
            if unicodedata.category(c) != 'Mn'
        ).lower()
    return value


def allowed_file(filename: str) -> bool:
    """Verifica se a extensão do arquivo é permitida."""
    ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx"}
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

from sqlalchemy import extract, cast, Integer

def aplicar_filtros_base(query, filtro_data, uvis_id):
    if filtro_data:
        try:
            # filtro_data = "2026-01"
            ano, mes = map(int, filtro_data.split('-'))
            
            # Forçamos a comparação de INTEIRO com INTEIRO
            query = query.filter(
                cast(extract('year', Solicitacao.data_agendamento), Integer) == ano,
                cast(extract('month', Solicitacao.data_agendamento), Integer) == mes
            )
            print(f"DEBUG SQL: Filtrando por Ano={ano} e Mes={mes}")
        except Exception as e:
            print(f"Erro no filtro de data: {e}")

    if uvis_id:
        query = query.filter(Solicitacao.usuario_id == int(uvis_id))
            
    return query

from functools import wraps
from flask import abort
from flask_login import current_user

def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                abort(401)
            if current_user.tipo_usuario not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

# --- DASHBOARD UVIS ---
@bp.route('/')
@login_required
def dashboard():
    if current_user.tipo_usuario == 'piloto':
        return redirect(url_for('main.piloto_os'))

    if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
        return redirect(url_for('main.admin_dashboard'))

    # ✅ UVIS: só as solicitações dela + carrega piloto para exibir
    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.piloto)
        )
        .filter(Solicitacao.usuario_id == current_user.id)
    )

    # Filtragem por status (original)
    filtro_status = request.args.get('status')
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    # Filtro por tipo de visita
    filtro_tipo_visita = request.args.get('tipo_visita')
    if filtro_tipo_visita:
        query = query.filter(Solicitacao.tipo_visita == filtro_tipo_visita)

    # Filtro por foco da ação
    filtro_foco = request.args.get('foco')
    if filtro_foco:
        query = query.filter(Solicitacao.foco == filtro_foco)  # Aqui, o valor de filtro_foco já é uma string


    # Paginação
    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(Solicitacao.data_criacao.desc())\
        .paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'dashboard.html',
        solicitacoes=paginacao.items,
        paginacao=paginacao,
    )


# --- PAINEL DE GESTÃO (Visualização para todos) ---
from flask_login import login_required, current_user
from datetime import datetime

@bp.route('/admin')
@login_required
def admin_dashboard():

    # 🔐 Controle de acesso
    if current_user.tipo_usuario not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Pode editar apenas admin e operario
    is_editable = current_user.tipo_usuario in ['admin', 'operario']

    # --- Captura filtros ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # --- Query base ---
    query = Solicitacao.query \
        .options(joinedload(Solicitacao.usuario)) \
        .join(Usuario)

    query = (
        Solicitacao.query
        .options(
            joinedload(Solicitacao.usuario),
            joinedload(Solicitacao.piloto)  # bom pra exibir o nome do piloto já atribuído
        )
        .join(Usuario)
    )

    pilotos = Pilotos.query.order_by(Pilotos.nome_piloto.asc()).all()


    # --- Aplicação dos filtros ---
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(
            Usuario.nome_uvis.ilike(f"%{filtro_unidade}%")
        )

    if filtro_regiao:
        query = query.filter(
            Usuario.regiao.ilike(f"%{filtro_regiao}%")
        )

    # Paginação
    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'admin.html',
        pedidos=paginacao.items,
        paginacao=paginacao,
        is_editable=is_editable,
        now=datetime.now(),
        pilotos=pilotos,
    )

@bp.route('/admin/exportar_excel')
@login_required
def exportar_excel():
    # 🔐 Permissão: somente admin e operario
    if current_user.tipo_usuario not in ['admin', 'operario']:
        flash('Permissão negada para exportar.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    try:
        filtro_status = request.args.get("status")
        filtro_unidade = request.args.get("unidade")
        filtro_regiao = request.args.get("regiao")

        # Evita Lazy Loading no Postgres
        query = (
            db.session.query(Solicitacao)
            .join(Usuario)
            .options(joinedload(Solicitacao.usuario))
        )

        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        if filtro_unidade:
            query = query.filter(
                Usuario.nome_uvis.ilike(f"%{filtro_unidade}%")
            )

        if filtro_regiao:
            query = query.filter(
                Usuario.regiao.ilike(f"%{filtro_regiao}%")
            )

        pedidos = query.order_by(
            Solicitacao.data_criacao.desc()
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Solicitações"

        headers = [
            "ID", "Unidade", "Região", "Data Agendada", "Hora",
            "Endereço Completo", "Latitude", "Longitude",
            "Foco", "Tipo Visita", "Altura", "Apoio CET?",
            "Observação", "Status", "Protocolo", "Justificativa"
        ]

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_num, p in enumerate(pedidos, 2):
            uvis_nome = p.usuario.nome_uvis if p.usuario else "Não informado"
            uvis_regiao = p.usuario.regiao if p.usuario else "Não informado"

            endereco_completo = (
                f"{p.logradouro or ''}, {p.numero or ''} - "
                f"{p.bairro or ''} - "
                f"{(p.cidade or '')}/{(p.uf or '')} - {p.cep or ''}"
            )
            if p.complemento:
                endereco_completo += f" - {p.complemento}"

            data_formatada = ""
            if p.data_agendamento:
                if isinstance(p.data_agendamento, (date, datetime)):
                    data_formatada = p.data_agendamento.strftime("%d/%m/%Y")
                else:
                    data_formatada = str(p.data_agendamento)

            row = [
                p.id,
                uvis_nome,
                uvis_regiao,
                data_formatada,
                str(p.hora_agendamento or ""),
                endereco_completo,
                p.latitude or "",
                p.longitude or "",
                p.foco,
                p.tipo_visita or "",
                p.altura_voo or "",
                "SIM" if p.apoio_cet else "NÃO",
                p.observacao or "",
                p.status,
                p.protocolo or "",
                p.justificativa or ""
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="relatorio_solicitacoes.xlsx",
            as_attachment=False,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO EXPORTAR EXCEL: {e}")
        flash(
            "Erro ao gerar o Excel. Verifique se os dados estão corretos.",
            "danger"
        )
        return redirect(url_for('main.admin_dashboard'))

from flask import request, jsonify, current_app, redirect, url_for, flash
import os, uuid
from werkzeug.utils import secure_filename

@bp.route('/admin/atualizar/<int:id>', methods=['POST'])
@login_required
def atualizar(id):

    # 🔐 Permissão
    if current_user.tipo_usuario not in ['admin', 'operario']:
        if request.accept_mimetypes.accept_html and not request.is_json:
            flash("Permissão negada.", "danger")
            return redirect(request.referrer or url_for("main.admin_dashboard"))
        return jsonify({"error": "Permissão negada"}), 403

    pedido = Solicitacao.query.get_or_404(id)

    # --- Atualização de campos ---
    pedido.protocolo = request.form.get('protocolo')
    pedido.status = request.form.get('status')
    pedido.justificativa = request.form.get('justificativa')
    pedido.latitude = request.form.get('latitude')
    pedido.longitude = request.form.get('longitude')

    # ✅ Atribuição de piloto (opcional)
    piloto_id = request.form.get("piloto_id")

    if piloto_id in (None, "", "null", "undefined"):
        pedido.piloto_id = None
    else:
        try:
            piloto_id_int = int(piloto_id)
            existe = Pilotos.query.get(piloto_id_int)
            if not existe:
                flash("Piloto selecionado não existe.", "warning")
                return redirect(request.referrer or url_for("main.admin_dashboard"))
            pedido.piloto_id = piloto_id_int
        except ValueError:
            flash("Piloto inválido.", "warning")
            return redirect(request.referrer or url_for("main.admin_dashboard"))

    # ✅ Regra de negócio: se aprovou, precisa ter piloto
    status_aprovacao = ["APROVADO", "APROVADO COM RECOMENDAÇÕES"]
    if pedido.status in status_aprovacao and not pedido.piloto_id:
        flash("Para aprovar, selecione um piloto responsável.", "warning")
        return redirect(request.referrer or url_for("main.admin_dashboard"))

    # Processamento de Anexo
    file = request.files.get("anexo")
    if file and file.filename:
        if allowed_file(file.filename):
            original_filename = secure_filename(file.filename)
            ext = original_filename.rsplit(".", 1)[1].lower()
            unique_name = f"sol_{pedido.id}_{uuid.uuid4().hex}.{ext}"
            upload_folder = get_upload_folder()
            file_path = os.path.join(upload_folder, unique_name)

            try:
                file.save(file_path)
                pedido.anexo_path = f"upload-files/{unique_name}"
                pedido.anexo_nome = original_filename
            except Exception as e:
                current_app.logger.error(f"Erro ao salvar arquivo físico: {e}")
                if request.accept_mimetypes.accept_html and not request.is_json:
                    flash("Falha ao salvar o arquivo no servidor.", "danger")
                    return redirect(request.referrer or url_for("main.admin_dashboard"))
                return jsonify({"error": "Falha ao salvar o arquivo no servidor."}), 500
        else:
            if request.accept_mimetypes.accept_html and not request.is_json:
                flash("Formato de arquivo não permitido.", "warning")
                return redirect(request.referrer or url_for("main.admin_dashboard"))
            return jsonify({"error": "Formato de arquivo não permitido."}), 400

    # commit final (MANTER APENAS ESTE BLOCO)
    try:
        db.session.commit()

        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json
        if is_ajax:
            return jsonify({
                "ok": True,
                "message": "Solicitação atualizada com sucesso!",
                "anexo_nome": pedido.anexo_nome,
                "piloto_id": pedido.piloto_id,
            }), 200

        flash("Solicitação atualizada com sucesso!", "success")
        return redirect(request.referrer or url_for("main.admin_dashboard"))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro de Banco (Atualizar ID {id}): {e}")

        if request.accept_mimetypes.accept_html and not request.is_json:
            flash("Erro ao gravar dados no banco de dados.", "danger")
            return redirect(request.referrer or url_for("main.admin_dashboard"))

        return jsonify({"error": "Erro ao gravar dados no banco de dados."}), 500


    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erro de Banco (Atualizar ID {id}): {e}")

        if request.accept_mimetypes.accept_html and not request.is_json:
            flash("Erro ao gravar dados no banco de dados.", "danger")
            return redirect(request.referrer or url_for("main.admin_dashboard"))

        return jsonify({"error": "Erro ao gravar dados no banco de dados."}), 500

    
# --- NOVO PEDIDO ---
from flask_login import login_required, current_user

@bp.route('/novo_cadastro', methods=['GET', 'POST'], endpoint='novo')
@login_required
def novo():

    hoje = date.today().isoformat()

    if request.method == 'POST':
        try:
            # --- Data ---
            data_str = request.form.get('data')
            hora_str = request.form.get('hora')

            data_obj = (
                datetime.strptime(data_str, '%Y-%m-%d').date()
                if data_str else None
            )

            hora_obj = (
                datetime.strptime(hora_str, '%H:%M').time()
                if hora_str else None
            )

            apoio_cet_bool = request.form.get('apoio_cet') == 'sim'

            nova_solicitacao = Solicitacao(
                data_agendamento=data_obj,
                hora_agendamento=hora_obj,

                cep=request.form.get('cep'),
                logradouro=request.form.get('logradouro'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                numero=request.form.get('numero'),
                uf=request.form.get('uf'),
                complemento=request.form.get('complemento'),

                foco=request.form.get('foco')
                tipo_visita=request.form.get('tipo_visita'),
                altura_voo=request.form.get('altura_voo'),
                apoio_cet=apoio_cet_bool,
                observacao=request.form.get('observacao'),

                latitude=request.form.get('latitude'),
                longitude=request.form.get('longitude'),

                # 🔑 RELAÇÃO CORRETA COM FLASK-LOGIN
                usuario_id=current_user.id,

                status='PENDENTE'
            )

            db.session.add(nova_solicitacao)
            db.session.commit()

            flash('Pedido enviado com sucesso!', 'success')
            return redirect(url_for('main.dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data ou hora.", "warning")

        except Exception as e:
            db.session.rollback()
            print(f"ERRO NOVO CADASTRO: {e}")
            flash("Erro ao salvar o pedido.", "danger")

    return render_template('cadastro.html', hoje=hoje)

# --- LOGIN ---
from flask_login import login_user

from flask_login import login_user, current_user

@bp.route('/login', methods=['GET', 'POST'])
def login():
    # Se já estiver logado, redireciona
    if current_user.is_authenticated:
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            return redirect(url_for('main.admin_dashboard'))
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        login_form = request.form.get('login')
        senha_form = request.form.get('senha')
        user = Usuario.query.filter_by(login=login_form).first()
        if user and user.check_senha(senha_form):
            login_user(user)  # 🔥 ÚNICO controle de login
            flash(
                f'Bem-vindo, {user.nome_uvis}! Login realizado com sucesso.',
                'success'
            )
            if user.tipo_usuario in ['admin', 'operario', 'visualizar']:
                return redirect(url_for('main.admin_dashboard'))
            return redirect(url_for('main.dashboard'))
        flash('Login ou senha incorretos. Tente novamente.', 'danger')

    return render_template('login.html')

# --- LOGOUT ---
from flask_login import logout_user, login_required

@bp.route('/logout')
@login_required
def logout():
    logout_user()          # 🔑 encerra o current_user
    session.clear()        # opcional (flash, tema, etc)
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('main.login'))


@bp.route("/forcar_erro")
def forcar_erro():
    1 / 0  # erro proposital
    return "nunca vai chegar aqui"

# Openpyxl (Excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                Spacer, Table, TableStyle)

# O objeto 'bp' precisa ser definido (Exemplo: bp = Blueprint('main', __name__))
# E 'Usuario' e 'Solicitacao' precisam ser seus modelos SQLAlchemy

# =======================================================================
# Função Auxiliar de Filtros (Reutilizada em todas as rotas)
# =======================================================================

from datetime import datetime

# =======================================================================
# ROTA 1: Visualização do Relatório (HTML)
# =======================================================================
from flask import redirect, render_template, request, session, url_for

from app import db
from app.models import Solicitacao, Usuario

@bp.route('/relatorios', methods=['GET'])
def relatorios():
    if not current_user.is_authenticated:
        return redirect(url_for('main.login'))

    try:
        # 1. Inicialize variáveis para evitar erro de 'not defined'
        uvis_disponiveis = []
        
        # 2. Busque as UVIS primeiro (se for admin)
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            uvis_disponiveis = (
                db.session.query(Usuario.id, Usuario.nome_uvis)
                .filter(Usuario.tipo_usuario == 'uvis')
                .order_by(Usuario.nome_uvis)
                .all()
            )

        # 3. Capture os parâmetros de Filtro
        mes_atual = request.args.get('mes', datetime.now().month, type=int)
        ano_atual = request.args.get('ano', datetime.now().year, type=int)
        uvis_id = request.args.get('uvis_id', type=int) if current_user.tipo_usuario != 'uvis' else current_user.id

        # 4. Monte a string de data para a função de filtro
        filtro_data = f"{ano_atual}-{mes_atual:02d}"

        # 5. Query base (AQUI USAMOS O EXTRACT AUTOMATICAMENTE ATRAVÉS DA FUNÇÃO)
        base_query = aplicar_filtros_base(
            db.session.query(Solicitacao),
            filtro_data,
            uvis_id
        )

        base_query = aplicar_filtros_base(db.session.query(Solicitacao), filtro_data, uvis_id)

        # ADICIONE ISSO AQUI:
        print("SQL EXECUTADO:", str(base_query.statement.compile(dialect=db.engine.dialect)))

        # =====================================================
        # 🔹 TOTAIS POR STATUS (JSON-safe)
        # =====================================================
        status_counts = {
            status: total
            for status, total in (
                base_query
                .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.status)
                .all()
            )
        }

        total_solicitacoes = sum(status_counts.values())
        total_aprovadas = status_counts.get("APROVADO", 0)
        total_aprovadas_com_recomendacoes = status_counts.get(
            "APROVADO COM RECOMENDAÇÕES", 0
        )
        total_recusadas = status_counts.get("NEGADO", 0)
        total_analise = status_counts.get("EM ANÁLISE", 0)
        total_pendentes = status_counts.get("PENDENTE", 0)

        # =====================================================
        # 🔹 FUNÇÃO GENÉRICA DE AGRUPAMENTO (JSON-safe)
        # =====================================================
        def agrupar_por(campo):
            resultados = (
                base_query
                .with_entities(campo, db.func.count(Solicitacao.id))
                .group_by(campo)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )

            return [
                (valor or "Não informado", total)
                for valor, total in resultados
            ]

        dados_status = agrupar_por(Solicitacao.status)
        dados_foco = agrupar_por(Solicitacao.foco)
        dados_tipo_visita = agrupar_por(Solicitacao.tipo_visita)
        dados_altura_voo = agrupar_por(Solicitacao.altura_voo)

        # =====================================================
        # 🔹 AGRUPAMENTOS COM JOIN (Corrigido para usar base_query)
        # =====================================================
        dados_regiao = [
            (regiao or "Não informado", total)
            for regiao, total in (
                base_query.join(Usuario)
                .with_entities(Usuario.regiao, db.func.count(Solicitacao.id))
                .group_by(Usuario.regiao)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )
        ]

        dados_unidade = [
            (uvis or "Não informado", total)
            for uvis, total in (
                base_query.join(Usuario)
                .filter(Usuario.tipo_usuario == 'uvis')
                .with_entities(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                .group_by(Usuario.nome_uvis)
                .order_by(db.func.count(Solicitacao.id).desc())
                .all()
            )
        ]

        # =====================================================
        # 🔹 HISTÓRICO MENSAL (Independente do filtro atual)
        # =====================================================
        dados_mensais = [
            (f"{int(ano_h):04d}-{int(mes_h):02d}", total)
            for ano_h, mes_h, total in (
                db.session.query(
                    extract('year', Solicitacao.data_agendamento),
                    extract('month', Solicitacao.data_agendamento),
                    db.func.count(Solicitacao.id)
                )
                .group_by(extract('year', Solicitacao.data_agendamento), extract('month', Solicitacao.data_agendamento))
                .order_by(extract('year', Solicitacao.data_agendamento), extract('month', Solicitacao.data_agendamento))
                .all()
            )
        ]

        anos_disponiveis = (
            sorted({m.split('-')[0] for m, _ in dados_mensais}, reverse=True)
            if dados_mensais else [ano_atual]
        )

        print(f"DEBUG FILTRO: Mês selecionado: {mes_atual} | String gerada: {filtro_data}")

        return render_template(
            'relatorios.html',
            total_solicitacoes=total_solicitacoes,
            total_aprovadas=total_aprovadas,
            total_aprovadas_com_recomendacoes=total_aprovadas_com_recomendacoes,
            total_recusadas=total_recusadas,
            total_analise=total_analise,
            total_pendentes=total_pendentes,
            dados_regiao=dados_regiao,
            dados_status=dados_status,
            dados_foco=dados_foco,
            dados_tipo_visita=dados_tipo_visita,
            dados_altura_voo=dados_altura_voo,
            dados_unidade=dados_unidade,
            dados_mensais=dados_mensais,
            mes_selecionado=mes_atual,
            ano_selecionado=ano_atual,
            anos_disponiveis=anos_disponiveis,
            uvis_id_selecionado=uvis_id,
            uvis_disponiveis=uvis_disponiveis,
            filtros={'total': total_solicitacoes}
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO NOS RELATÓRIOS: {e}")
        return render_template(
            "erro.html",
            codigo=500,
            titulo="Erro nos Relatórios",
            mensagem="Houve um erro técnico ao processar os dados."
        )
    

import os
import tempfile
from datetime import datetime
from io import BytesIO

from flask import send_file, request
from flask_login import login_required, current_user
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

@bp.route('/admin/exportar_relatorio_pdf')
@login_required
def exportar_relatorio_pdf():
    # -------------------------
    # 1. Parâmetros e filtros (IGUAL ao /relatorios)
    # -------------------------
    mes = int(request.args.get('mes', datetime.now().month))
    ano = int(request.args.get('ano', datetime.now().year))
    orient = request.args.get('orient', default='portrait')  # 'portrait' ou 'landscape'
    filtro_data = f"{ano}-{mes:02d}"

    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Query base e detalhe
    # -------------------------
    base_query = aplicar_filtros_base(
        db.session.query(Solicitacao),
        filtro_data,
        uvis_id
    )

    query_detalhe = aplicar_filtros_base(
        db.session.query(Solicitacao, Usuario).join(Usuario, Usuario.id == Solicitacao.usuario_id),
        filtro_data,
        uvis_id
    )

    query_results = query_detalhe.order_by(Solicitacao.data_criacao.desc()).all()

    # -------------------------
    # 3. Totais
    # -------------------------
    total_solicitacoes = base_query.count()
    total_aprovadas = base_query.filter(Solicitacao.status == "APROVADO").count()
    total_aprovadas_com_recomendacoes = base_query.filter(
        Solicitacao.status == "APROVADO COM RECOMENDAÇÕES"
    ).count()
    total_recusadas = base_query.filter(Solicitacao.status == "NEGADO").count()
    total_analise = base_query.filter(Solicitacao.status == "EM ANÁLISE").count()
    total_pendentes = base_query.filter(Solicitacao.status == "PENDENTE").count()

    STATUS_COLORS = {
        "APROVADO": "#2ecc71",
        "APROVADO COM RECOMENDAÇÕES": "#ee650a",
        "EM ANÁLISE": "#f1c40f",
        "PENDENTE": "#3498db",
        "NEGADO": "#e74c3c",
    }

    # -------------------------
    # 4. Agrupamentos
    # -------------------------
    dados_regiao = [
        (regiao or "Não informado", total)
        for regiao, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)).join(Usuario),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.regiao)
            .all()
        )
    ]

    dados_status = [
        (status or "Não informado", total)
        for status, total in (
            base_query
            .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.status)
            .all()
        )
    ]

    dados_foco = [
        (foco or "Não informado", total)
        for foco, total in (
            base_query
            .with_entities(Solicitacao.foco, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.foco)
            .all()
        )
    ]

    dados_tipo_visita = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_visita)
            .all()
        )
    ]

    dados_altura_voo = [
        (altura or "Não informado", total)
        for altura, total in (
            base_query
            .with_entities(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.altura_voo)
            .all()
        )
    ]

    dados_unidade = [
        (uvis_nome or "Não informado", total)
        for uvis_nome, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                .join(Usuario)
                .filter(Usuario.tipo_usuario == 'uvis'),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.nome_uvis)
            .all()
        )
    ]

    if db.engine.name == 'postgresql':
        func_mes = db.func.to_char(Solicitacao.data_agendamento, 'YYYY-MM')
    else:
        func_mes = db.func.strftime('%Y-%m', Solicitacao.data_agendamento)

    dados_mensais = [
        tuple(row) for row in (
            db.session.query(func_mes.label('mes'), db.func.count(Solicitacao.id))
            .group_by('mes')
            .order_by('mes')
            .all()
        )
    ]

    # -------------------------
    # 5. Preparar PDF
    # -------------------------
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    caminho_pdf = tmp_pdf.name
    tmp_pdf.close()

    pagesize = landscape(A4) if orient == 'landscape' else A4

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=pagesize,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=16*mm, bottomMargin=16*mm
    )

    styles = getSampleStyleSheet()

    # Tipografia melhor
    title_style = ParagraphStyle(
        'title',
        parent=styles['Title'],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'subtitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor('#555'),
        spaceAfter=12
    )

    section_h = ParagraphStyle(
        'sec',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#0d6efd'),
        spaceBefore=10,
        spaceAfter=6
    )

    normal = ParagraphStyle(
        'normal',
        parent=styles['Normal'],
        fontSize=9.5,
        leading=13
    )

    cell_style = ParagraphStyle(
        'cell',
        parent=styles['BodyText'],
        fontSize=8.6,
        leading=11,
        textColor=colors.HexColor('#222'),
        wordWrap='CJK',
        splitLongWords=True
    )

    story = []

    # -------------------------
    # CAPA (Resumo)
    # -------------------------
    story.append(Paragraph(f"Relatório Mensal — {mes:02d}/{ano}", title_style))

    filtro_txt = f"Filtro: {filtro_data}"
    if uvis_id:
        filtro_txt += f" | UVIS ID: {uvis_id}"
    else:
        filtro_txt += " | UVIS: Todas"
    story.append(Paragraph(filtro_txt, subtitle_style))

    # Cards do resumo (bem mais bonito)
    def resumo_cards():
        cards = [
            ("Total", total_solicitacoes, '#0d6efd'),
            ("Aprovadas", total_aprovadas, '#198754'),
            ("Aprov. c/ Recom.", total_aprovadas_com_recomendacoes, '#6c757d'),
            ("Negadas", total_recusadas, '#dc3545'),
            ("Em Análise", total_analise, '#ffc107'),
            ("Pendentes", total_pendentes, '#0dcaf0'),
        ]

        rows = []
        row = []
        for i, (label, value, hexcolor) in enumerate(cards, start=1):
            box = Table(
                [
                    [Paragraph(label, ParagraphStyle('l', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666')))],
                    [Paragraph(str(value), ParagraphStyle('v', parent=styles['Normal'], fontSize=18, leading=20, textColor=colors.HexColor(hexcolor)))]
                ],
                colWidths=[48*mm] if orient == 'portrait' else [52*mm],
            )
            box.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8f9fa')),
                ('BOX', (0,0), (-1,-1), 0.6, colors.HexColor('#e5e7eb')),
                ('LEFTPADDING', (0,0), (-1,-1), 8),
                ('RIGHTPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))

            row.append(box)
            if len(row) == 3:
                rows.append(row)
                row = []

        if row:
            # completa a linha
            while len(row) < 3:
                row.append(Spacer(1, 1))
            rows.append(row)

        grid = Table(rows, colWidths=None)
        grid.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        return grid

    story.append(resumo_cards())
    story.append(Spacer(1, 10))

    # -------------------------
    # TABELAS (DADOS ESCRITOS PRIMEIRO)
    # -------------------------
    def add_count_table(titulo, dados, col1="Categoria"):
        story.append(Paragraph(titulo, section_h))

        rows = [
            [Paragraph(col1, ParagraphStyle('th', parent=cell_style, textColor=colors.white, fontSize=9)),
             Paragraph("Total", ParagraphStyle('th2', parent=cell_style, textColor=colors.white, fontSize=9))]
        ]

        for nome, total in (dados or [("Nenhum", 0)]):
            rows.append([Paragraph(str(nome), cell_style), Paragraph(str(total), cell_style)])

        tbl = Table(rows, repeatRows=1, colWidths=[140*mm, 25*mm] if orient == 'portrait' else [190*mm, 30*mm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#0d6efd')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),9),
            ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#fbfdff')]),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

    # Um “Resumo por agrupamento” em sequência (mais agradável)
    story.append(Paragraph("Resumo por Agrupamentos", section_h))
    story.append(Paragraph("Abaixo estão os agrupamentos do mês selecionado, apresentados em formato de tabela.", normal))
    story.append(Spacer(1, 6))

    add_count_table("Agrupamento — Região", dados_regiao)
    add_count_table("Agrupamento — Status", dados_status)
    add_count_table("Agrupamento — Foco", dados_foco)
    add_count_table("Agrupamento — Tipo de Visita", dados_tipo_visita)
    add_count_table("Agrupamento — Altura do Voo", dados_altura_voo)
    add_count_table("Agrupamento — Unidade (UVIS)", dados_unidade)
    add_count_table("Histórico Mensal (tabela)", dados_mensais, col1="Mês")

    # -------------------------
    # ✅ GRÁFICOS (AGORA DEPOIS DOS DADOS ESCRITOS)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Gráficos", section_h))
    story.append(Paragraph("Os gráficos abaixo representam visualmente os dados apresentados nas tabelas anteriores.", normal))
    story.append(Spacer(1, 8))

    def safe_img_from_plt(fig, width_mm=170):
        bio = BytesIO()
        fig.tight_layout()
        fig.savefig(bio, format='png', dpi=220, bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return RLImage(bio, width=width_mm*mm)

    if MATPLOTLIB_AVAILABLE:
        try:
            # 1) Donut por status (mais limpo)
            labels = [s for s, _ in dados_status]
            values = [c for _, c in dados_status]
            colors_status = [STATUS_COLORS.get(s, "#bdc3c7") for s in labels]

            fig1, ax1 = plt.subplots(figsize=(6.4, 3.0))
            def autopct(p): return f'{p:.0f}%' if p >= 6 else ''
            wedges, *_ = ax1.pie(
                values or [1],
                labels=None,
                colors=colors_status,
                autopct=autopct,
                startangle=90,
                pctdistance=0.78,
                textprops={'fontsize': 9}
            )
            centre_circle = plt.Circle((0, 0), 0.58, fc='white')
            ax1.add_artist(centre_circle)
            ax1.legend(wedges, labels, loc='center left', bbox_to_anchor=(1.02, 0.5),
                       fontsize=9, frameon=False)
            ax1.set_title('Distribuição por Status', fontsize=11, pad=10)
            ax1.axis('equal')

            story.append(safe_img_from_plt(fig1, width_mm=170))
            story.append(Spacer(1, 10))

            # 2) Top UVIS (barra horizontal)
            u_names = [u for u, _ in dados_unidade[:10]]
            u_vals = [c for _, c in dados_unidade[:10]]

            fig2, ax2 = plt.subplots(figsize=(7.2, 3.0))
            ax2.barh(u_names[::-1] or ['Nenhum'], u_vals[::-1] or [0])
            ax2.set_xlabel('Total', fontsize=9)
            ax2.set_title('Top UVIS', fontsize=11, pad=10)
            ax2.tick_params(axis='both', labelsize=9)
            ax2.grid(axis='x', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig2, width_mm=180 if orient == 'landscape' else 170))
            story.append(Spacer(1, 10))

            # 3) Histórico mensal (linha)
            months = [m for m, _ in dados_mensais]
            counts = [c for _, c in dados_mensais]

            fig3, ax3 = plt.subplots(figsize=(7.2, 3.0))
            if months:
                ax3.plot(range(len(months)), counts, marker='o', linewidth=1.6)
                ax3.set_xticks(range(len(months)))
                ax3.set_xticklabels(months, rotation=45, ha='right', fontsize=9)
            ax3.set_title('Histórico Mensal', fontsize=11, pad=10)
            ax3.tick_params(axis='y', labelsize=9)
            ax3.grid(axis='y', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig3, width_mm=185 if orient == 'landscape' else 170))
            story.append(Spacer(1, 8))

        except Exception:
            story.append(Paragraph("Gráficos indisponíveis (erro ao gerar).", normal))
    else:
        story.append(Paragraph("Matplotlib não disponível — gráficos foram omitidos.", normal))

        # -------------------------
    # DETALHES (Registros Detalhados)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Registros Detalhados", section_h))
    story.append(Paragraph("Listagem completa dos registros retornados pelo filtro selecionado.", normal))
    story.append(Spacer(1, 8))

    registros_header = [
        'Data', 'Hora', 'Unidade', 'Região', 'Protocolo',
        'Status', 'Foco', 'Tipo Visita', 'Altura Voo', 'Observação'
    ]

    hdr_style = ParagraphStyle(
        'hdr',
        parent=cell_style,
        textColor=colors.white,
        fontSize=7.8,
        leading=9.2
    )

    cell_style_small = ParagraphStyle(
        'cell_small',
        parent=cell_style,
        fontSize=7.6,
        leading=9.2,
        wordWrap='CJK',
        splitLongWords=True
    )

    registros_rows = [[Paragraph(h, hdr_style) for h in registros_header]]

    for s, u in query_results:
        data_str = s.data_criacao.strftime("%d/%m/%Y") if getattr(s, 'data_criacao', None) else ''
        hora_str = getattr(s, 'hora_agendamento', '')
        hora_str = hora_str.strftime("%H:%M") if hasattr(hora_str, 'strftime') else str(hora_str or '')

        unidade = getattr(u, 'nome_uvis', '') or "Não informado"
        regiao = getattr(u, 'regiao', '') or "Não informado"

        protocolo = getattr(s, 'protocolo', '') or ''
        status = getattr(s, 'status', '') or ''
        foco = getattr(s, 'foco', '') or ''
        tipo_visita = getattr(s, 'tipo_visita', '') or ''
        altura_voo = getattr(s, 'altura_voo', '') or ''
        obs = getattr(s, 'observacao', '') or ''

        registros_rows.append([
            Paragraph(str(data_str), cell_style_small),
            Paragraph(str(hora_str), cell_style_small),
            Paragraph(str(unidade), cell_style_small),
            Paragraph(str(regiao), cell_style_small),
            Paragraph(str(protocolo), cell_style_small),
            Paragraph(str(status), cell_style_small),
            Paragraph(str(foco), cell_style_small),
            Paragraph(str(tipo_visita), cell_style_small),
            Paragraph(str(altura_voo), cell_style_small),
            Paragraph(str(obs), cell_style_small),
        ])

    # ✅ Larguras base (as suas), mas vamos “encaixar” no doc.width automaticamente
    base_col_widths = [
        18*mm, 14*mm, 28*mm, 22*mm, 22*mm,
        22*mm, 22*mm, 26*mm, 18*mm, 60*mm
    ]

    # ✅ Se a soma estourar a largura útil da página, escala proporcionalmente
    total_w = sum(base_col_widths)
    max_w = doc.width  # largura útil = página - margens

    if total_w > max_w:
        scale = max_w / total_w
        colWidths = [w * scale for w in base_col_widths]
    else:
        colWidths = base_col_widths

    # ✅ Quantidade de linhas por página (ajuste fino)
    chunk_size = 28 if orient == 'landscape' else 24

    # 🔥 renderiza em blocos para não ficar pesado e manter header repetido
    for i in range(0, len(registros_rows), chunk_size):
        chunk = registros_rows[i:i + chunk_size]

        tbl = Table(
            chunk,
            repeatRows=1,
            colWidths=colWidths,
            hAlign='LEFT'  # ✅ evita “puxar” pro centro e cortar laterais
        )

        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fbfdff')]),

            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (-1,0), 'LEFT'),
            ('ALIGN', (0,1), (-1,-1), 'LEFT'),

            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),

            # ✅ reforço de quebra de linha dentro das células
            ('WORDWRAP', (0,0), (-1,-1), 'CJK'),
        ]))

        story.append(tbl)
        story.append(Spacer(1, 6))

        if i + chunk_size < len(registros_rows):
            story.append(PageBreak())


    # -------------------------
    # Header/Footer
    # -------------------------
    def _header_footer(canvas, doc_):
        canvas.saveState()
        w, h = pagesize

        canvas.setFillColor(colors.HexColor('#0d6efd'))
        canvas.rect(doc_.leftMargin, h-(12*mm), doc_.width, 3, fill=1, stroke=0)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor('#777'))
        canvas.drawString(doc_.leftMargin, 9*mm, f"Relatório — {mes:02d}/{ano} — IJASystem")
        canvas.drawRightString(doc_.leftMargin + doc_.width, 9*mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    nome_arquivo = f"relatorio_IJASystem_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    return send_file(
        caminho_pdf,
        as_attachment=True,
        download_name=f"{nome_arquivo}.pdf",
        mimetype="application/pdf"
    )


# =======================================================================
# ROTA 3: Exportar Excel (Com Filtro UVIS) - Layout “bonito” igual Excel
# =======================================================================
@bp.route('/admin/exportar_relatorio_excel')
@login_required
def exportar_relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # -------------------------
    # 1. Parâmetros e filtros
    # -------------------------
    mes = request.args.get('mes', datetime.now().month, type=int)
    ano = request.args.get('ano', datetime.now().year, type=int)
    filtro_data = f"{ano}-{mes:02d}"

    # Controle de acesso UVIS
    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Busca de Dados
    # -------------------------
    query_dados = db.session.query(
        Solicitacao.id,
        Solicitacao.status,
        Solicitacao.foco,
        Solicitacao.tipo_visita,
        Solicitacao.altura_voo,
        Solicitacao.data_agendamento,
        Solicitacao.hora_agendamento,

        # Endereço (campos separados no banco)
        Solicitacao.cep,
        Solicitacao.logradouro,
        Solicitacao.numero,
        Solicitacao.bairro,
        Solicitacao.cidade,
        Solicitacao.uf,

        Solicitacao.latitude,
        Solicitacao.longitude,

        # UVIS
        Usuario.nome_uvis,
        Usuario.regiao
    ).join(Usuario, Usuario.id == Solicitacao.usuario_id)

   # Filtro de data pelo AGENDAMENTO (mês/ano)
    if db.engine.name == 'postgresql':
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.to_char(Solicitacao.data_agendamento, 'YYYY-MM') == filtro_data
        )
    else:
        query_dados = query_dados.filter(
            Solicitacao.data_agendamento.isnot(None),
            db.func.strftime('%Y-%m', Solicitacao.data_agendamento) == filtro_data
        )
    # Filtro opcional por UVIS
    if uvis_id:
        query_dados = query_dados.filter(Solicitacao.usuario_id == uvis_id)

  # Ordenar pelo agendamento (e hora como critério secundário)
    dados = query_dados.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc()
    ).all()

    # Se tiver filtro UVIS, pega o nome pra ajudar no nome do arquivo
    nome_uvis_filtro = None
    if uvis_id:
        nome_uvis_filtro = db.session.query(Usuario.nome_uvis).filter(Usuario.id == uvis_id).scalar()

    # -------------------------
    # 3. Helper: montar endereço em 1 LINHA (compacto igual Excel)
    # -------------------------
    def montar_endereco(row):
        partes_rua = []
        if row.logradouro:
            partes_rua.append(row.logradouro.strip())
        if row.numero is not None and str(row.numero).strip():
            partes_rua.append(str(row.numero).strip())

        rua_numero = ", ".join([p for p in partes_rua if p]).strip()

        cidade_uf = ""
        if row.cidade and row.uf:
            cidade_uf = f"{row.cidade.strip()}/{row.uf.strip()}"
        elif row.cidade:
            cidade_uf = row.cidade.strip()
        elif row.uf:
            cidade_uf = row.uf.strip()

        bairro_cidade = " - ".join([p for p in [(row.bairro or "").strip(), cidade_uf] if p]).strip()
        cep_txt = f"CEP {row.cep.strip()}" if row.cep else ""

        # Formato final: "Rua, 123 | Bairro - Cidade/UF | CEP 00000-000"
        return " | ".join([p for p in [rua_numero, bairro_cidade, cep_txt] if p])

    # -------------------------
    # 4. Criar arquivo Excel
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # ✅ UVIS vem no começo agora
    colunas = [
        "UVIS", "Região",
        "ID", "Status", "Foco", "Tipo Visita", "Altura Voo",
        "Data Agendamento", "Hora Agendamento",
        "ENDEREÇO DE AÇÃO",
        "Latitude", "Longitude"
    ]

    # Estilos (bem padrão Excel)
    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra1 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")

    # ✅ alinhamento igual ao print (compacto e central vertical)
    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")

    # Cabeçalho
    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Altura do cabeçalho (compacto)
    ws.row_dimensions[1].height = 22

    # Preenchimento de linhas
    for row_num, row in enumerate(dados, 2):
        data_agendamento_fmt = row.data_agendamento.strftime("%d/%m/%Y") if row.data_agendamento else ""
        hora_agendamento_fmt = row.hora_agendamento.strftime("%H:%M") if row.hora_agendamento else ""

        endereco_acao = montar_endereco(row)

        values = [
            row.nome_uvis,
            row.regiao,
            row.id,
            row.status,
            row.foco,
            row.tipo_visita,
            row.altura_voo,
            data_agendamento_fmt,
            hora_agendamento_fmt,
            endereco_acao,
            row.latitude,
            row.longitude
        ]

        # Altura das linhas (igual Excel “padrão bonito”)
        ws.row_dimensions[row_num].height = 20

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.border = thin_border
            cell.fill = zebra1 if (row_num % 2 == 0) else zebra2

            # Centraliza campos curtos, texto fica alinhado à esquerda (igual print)
            if col_index in (3, 7, 8, 9, 11, 12):  # ID, Altura, Data, Hora, Lat, Long
                cell.alignment = center
            else:
                cell.alignment = left_center

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Auto-filtro no cabeçalho
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    # ✅ Larguras “na mão” (fica igual ao print / bem organizado)
    larguras = {
        "A": 24,  # UVIS
        "B": 12,  # Região
        "C": 6,   # ID
        "D": 18,  # Status
        "E": 22,  # Foco
        "F": 16,  # Tipo Visita
        "G": 10,  # Altura Voo
        "H": 14,  # Data
        "I": 14,  # Hora
        "J": 90,  # ENDEREÇO DE AÇÃO
        "K": 14,  # Latitude
        "L": 14   # Longitude
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    # -------------------------
    # 5. Gerar arquivo em memória
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo
    nome_arquivo = f"relatorio_IJASystem_{ano}_{mes:02d}"
    if uvis_id:
        safe_nome = (nome_uvis_filtro or f"ID_{uvis_id}").replace(" ", "_")
        nome_arquivo += f"_UVIS_{safe_nome}"

    return send_file(
        output,
        download_name=f"{nome_arquivo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




from flask import flash, redirect, url_for, render_template, request
from flask_login import login_required, current_user
from app import db
from app.models import Solicitacao, Usuario
from datetime import datetime
from sqlalchemy.orm import joinedload

@bp.route('/admin/editar_completo/<int:id>', methods=['GET', 'POST'])
@login_required
def admin_editar_completo(id):
    # 🔐 Controle de acesso
    if current_user.tipo_usuario != 'admin':
        flash('Permissão negada. Apenas administradores podem acessar esta página.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Busca segura com joinedload para evitar lazy-loading
    pedido = Solicitacao.query.options(joinedload(Solicitacao.usuario)).get_or_404(id)

    # Listas para selects do template (pré-preenchimento)
    status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]
    foco_opcoes = ["Foco 1", "Foco 2", "Foco 3"]  # ajuste conforme seus valores reais
    tipo_visita_opcoes = ["Tipo 1", "Tipo 2", "Tipo 3"]  # ajuste conforme seus valores reais
    uf_opcoes = ["AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
                 "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"]

    if request.method == 'POST':
        try:
            # Guardar estado anterior de data/hora
            antes_data = pedido.data_agendamento
            antes_hora = pedido.hora_agendamento

            # 1️⃣ Atualizar datas e horas
            data_str = request.form.get('data_agendamento')
            hora_str = request.form.get('hora_agendamento')

            pedido.data_agendamento = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            pedido.hora_agendamento = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None

            # 2️⃣ Atualizar campos principais
            pedido.foco = request.form.get('foco') or pedido.foco
            pedido.tipo_visita = request.form.get('tipo_visita') or pedido.tipo_visita
            pedido.altura_voo = request.form.get('altura_voo') or pedido.altura_voo
            pedido.apoio_cet = request.form.get('apoio_cet', 'não').lower() == 'sim'
            pedido.observacao = request.form.get('observacao') or pedido.observacao

            # 3️⃣ Atualizar endereço
            pedido.cep = request.form.get('cep') or pedido.cep
            pedido.logradouro = request.form.get('logradouro') or pedido.logradouro
            pedido.numero = request.form.get('numero') or pedido.numero
            pedido.bairro = request.form.get('bairro') or pedido.bairro
            pedido.cidade = request.form.get('cidade') or pedido.cidade
            pedido.uf = request.form.get('uf') or pedido.uf
            pedido.complemento = request.form.get('complemento') or pedido.complemento

            # 4️⃣ Atualizar protocolo, status, justificativa e coordenadas
            pedido.protocolo = request.form.get('protocolo') or pedido.protocolo
            pedido.status = request.form.get('status') or pedido.status
            pedido.justificativa = request.form.get('justificativa') or pedido.justificativa

            lat = request.form.get('latitude')
            lon = request.form.get('longitude')
            pedido.latitude = float(lat) if lat else None
            pedido.longitude = float(lon) if lon else None

            # Commit
            db.session.commit()

            # 🔔 Notificação se agendamento mudou
            mudou_agendamento = (antes_data != pedido.data_agendamento) or (antes_hora != pedido.hora_agendamento)
            if pedido.data_agendamento and mudou_agendamento:
                data_fmt = pedido.data_agendamento.strftime("%d/%m/%Y")
                hora_fmt = pedido.hora_agendamento.strftime("%H:%M") if pedido.hora_agendamento else "00:00"
                criar_notificacao(
                    usuario_id=pedido.usuario_id,
                    titulo="Agendamento atualizado",
                    mensagem=f"Sua solicitação foi agendada para {data_fmt} às {hora_fmt}.",
                    link=url_for("main.agenda")
                )

            flash('Solicitação atualizada com sucesso!', 'success')
            return redirect(url_for('main.admin_dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data/hora: {ve}", 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar a solicitação: {e}", 'danger')

    return render_template(
        'admin_editar_completo.html',
        pedido=pedido,
        status_opcoes=status_opcoes,
        foco_opcoes=foco_opcoes,
        tipo_visita_opcoes=tipo_visita_opcoes,
        uf_opcoes=uf_opcoes
    )

from flask_login import current_user, login_required

@bp.route('/admin/deletar/<int:id>', methods=['POST'], endpoint='deletar_registro')
@login_required
def deletar(id):
    # Verifica se é admin
    if current_user.tipo_usuario != 'admin':  # <-- CORRETO: tipo_usuario
        flash('Permissão negada. Apenas administradores podem deletar registros.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Busca a solicitação
    pedido = Solicitacao.query.get_or_404(id)
    pedido_id = pedido.id

    # Nome do autor da solicitação
    autor_nome = pedido.usuario.nome_uvis if pedido.usuario else "UVIS"

    try:
        db.session.delete(pedido)
        db.session.commit()
    except Exception:
        db.session.rollback()
        # Não mostra erro ao usuário

    flash(f"Pedido #{pedido_id} da {autor_nome} deletado permanentemente.", "success")
    return redirect(url_for('main.admin_dashboard'))


from flask_login import login_required, current_user

from flask_login import login_required, current_user
import traceback

from flask import request, render_template
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from datetime import datetime
import json

@bp.route("/agenda")
@login_required
def agenda():
    try:
        # --- Usuário atual ---
        user_tipo = current_user.tipo_usuario
        user_id = current_user.id

        # --- Filtros GET ---
        filtro_status = request.args.get("status") or None
        filtro_uvis_id = request.args.get("uvis_id", type=int)
        mes = request.args.get("mes", datetime.now().month, type=int)
        ano = request.args.get("ano", datetime.now().year, type=int)
        d = request.args.get("d")
        initial_date = d or f"{ano}-{mes:02d}-01"

        # --- Query base ---
        query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

        if user_tipo not in ["admin", "operario", "visualizar"]:
            query = query.filter(Solicitacao.usuario_id == user_id)
            filtro_uvis_id = None
            pode_filtrar_uvis = False
        else:
            pode_filtrar_uvis = True
            if filtro_uvis_id:
                query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == 'postgresql':
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

        eventos = query.all()

        # --- Monta eventos para o FullCalendar ---
        agenda_eventos = []
        for e in eventos:
            try:
                data = e.data_agendamento.strftime("%Y-%m-%d")
                hora = e.hora_agendamento.strftime("%H:%M") if e.hora_agendamento else "00:00"
                uvis_nome = e.usuario.nome_uvis if e.usuario else "UVIS"
            except:
                uvis_nome = "UVIS"

            ev = {
                "id": str(e.id),
                "title": f"{e.foco} - {uvis_nome}",
                "start": f"{data}T{hora}",
                "color": (
                    "#198754" if e.status == "APROVADO" else
                    "#ffa023" if e.status == "APROVADO COM RECOMENDAÇÕES" else
                    "#dc3545" if e.status == "NEGADO" else
                    "#e9fa05" if e.status == "EM ANÁLISE" else
                    "#0d6efd"
                ),
                "extendedProps": {
                    "foco": e.foco,
                    "uvis": uvis_nome,
                    "hora": hora,
                    "status": e.status
                }
            }
            agenda_eventos.append(ev)

        # --- Variáveis para filtros ---
        status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]

        uvis_disponiveis = []
        if user_tipo in ["admin", "operario", "visualizar"]:
            uvis_disponiveis = db.session.query(Usuario.id, Usuario.nome_uvis).filter(Usuario.tipo_usuario == "uvis").order_by(Usuario.nome_uvis).all()

        # --- Anos disponíveis ---
        if db.engine.name == 'postgresql':
            func_ano = db.func.to_char(Solicitacao.data_agendamento, "YYYY")
        else:
            func_ano = db.func.strftime("%Y", Solicitacao.data_agendamento)

        anos_raw = db.session.query(func_ano).filter(Solicitacao.data_agendamento.isnot(None)).distinct().order_by(func_ano.desc()).all()
        anos_disponiveis = [int(a[0]) for a in anos_raw if a and a[0]]
        if not anos_disponiveis:
            anos_disponiveis = [datetime.now().year]

        # --- Dicionário de filtros para template ---
        filtros = {
            "uvis_id": filtro_uvis_id,
            "status": filtro_status,
            "mes": mes,
            "ano": ano
        }

        return render_template(
            "agenda.html",
            eventos_json=json.dumps(agenda_eventos),
            filtros=filtros,
            status_opcoes=status_opcoes,
            uvis_disponiveis=uvis_disponiveis,
            anos_disponiveis=anos_disponiveis,
            initial_date=initial_date,
            pode_filtrar_uvis=pode_filtrar_uvis
        )

    except Exception as e:
        import traceback
        print("TRACEBACK COMPLETO:")
        traceback.print_exc()
        return f"ERRO NA AGENDA: {str(e)}"


@bp.route("/agenda/exportar_excel", endpoint="agenda_exportar_excel")
@login_required
def exportar_agenda_excel():  
    if current_user.tipo_usuario != "admin":
        abort(403)  # Forbidden

    user_tipo = current_user.tipo_usuario
    user_id = current_user.id
    export_all = request.args.get("all") == "1"

    # filtros
    filtro_status = None if export_all else (request.args.get("status") or None)
    filtro_uvis_id = None if export_all else request.args.get("uvis_id", type=int)
    mes = None if export_all else request.args.get("mes", type=int)
    ano = None if export_all else request.args.get("ano", type=int)

    query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

    if filtro_uvis_id:
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)
    if mes and ano:
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == 'postgresql':
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

    query = query.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc()
    )
    eventos = query.all()
    # -----------------------------
    # Monta XLSX
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    headers = [
        "DATA",
        "HORÁRIO",
        "REGIÃO",
        "UVIS",
        "CET",
        "ENDEREÇO DA AÇÃO",
        "CEP",
        "FOCO DA AÇÃO",
        "COORDENADA GEOGRÁFICA",
        "Altura dos Voos",
        "Protocolo DECA",
        "Status",
    ]
    ws.append(headers)

    for p in eventos:
        endereco_completo = (
            f"{p.logradouro or ''}, {getattr(p, 'numero', '')} - "
            f"{p.bairro or ''} - "
            f"{(p.cidade or '')}/{(p.uf or '')} - "
            f"{p.cep or ''}"
        )
        if getattr(p, "complemento", None):
            endereco_completo += f" - {p.complemento}"

        cet_txt = "SIM" if getattr(p, "apoio_cet", None) else "NÃO"
        data_str = p.data_agendamento.strftime("%d/%m/%Y") if p.data_agendamento else ""
        hora_str = p.hora_agendamento.strftime("%H:%M") if p.hora_agendamento else ""
        uvis_nome = p.usuario.nome_uvis if getattr(p, "usuario", None) else ""
        regiao = p.usuario.regiao if getattr(p, "usuario", None) else ""
        lat = getattr(p, "latitude", "") or ""
        lon = getattr(p, "longitude", "") or ""
        coordenada = f"{lat},{lon}" if (lat or lon) else ""
        protocolo_deca = getattr(p, "protocolo_deca", None) or getattr(p, "protocolo", "") or ""

        ws.append([
            data_str,
            hora_str,
            regiao,
            uvis_nome,
            cet_txt,
            endereco_completo,
            getattr(p, "cep", "") or "",
            getattr(p, "foco", "") or "",
            coordenada,
            getattr(p, "altura_voo", "") or "",
            protocolo_deca,
            getattr(p, "status", "") or "",
        ])

    # -----------------------------
    # Estilo
    # -----------------------------
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap if cell.row > 1 else center

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = "agenda_tudo.xlsx" if export_all else "agenda_exportada.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# =================================================
# NOTIFICAÇÕES (Flask-Login: login_required + current_user)
# Requer no topo:
# from flask_login import login_required, current_user
# from flask import redirect, url_for, render_template
# from datetime import datetime, date
# from sqlalchemy.orm import joinedload
# from zoneinfo import ZoneInfo
# =================================================

from zoneinfo import ZoneInfo

TZ_BR = ZoneInfo("America/Sao_Paulo")

def agora_brasilia_naive():
    """
    Retorna datetime no horário de Brasília, mas sem tzinfo (naive),
    para bater com db.DateTime (sem timezone).
    """
    return datetime.now(TZ_BR).replace(tzinfo=None)


# -------------------------------------------------
# CRIAR NOTIFICAÇÃO
# -------------------------------------------------
def criar_notificacao(usuario_id, titulo, mensagem="", link=None):
    n = Notificacao(
        usuario_id=usuario_id,
        titulo=titulo,
        mensagem=mensagem or "",
        link=link,
        criada_em=agora_brasilia_naive(),  # ✅ Brasília
    )
    db.session.add(n)
    db.session.commit()
    return n


# -------------------------------------------------
# GARANTIR NOTIFICAÇÕES DO DIA (sem duplicar)
# ✅ REGRA: se já existiu (mesmo apagada), NÃO recria
# -------------------------------------------------
def garantir_notificacoes_do_dia(usuario_id):
    hoje = date.today()

    ags = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter_by(usuario_id=usuario_id)
        .filter(Solicitacao.data_agendamento == hoje)
        .all()
    )

    for s in ags:
        hora_fmt = s.hora_agendamento.strftime("%H:%M") if s.hora_agendamento else "00:00"

        # 🔒 chave estável (muda por dia por conta do d=hoje)
        link = url_for("main.agenda", sid=s.id, d=hoje.isoformat())

        # ✅ Se já existe (inclusive apagada), NÃO cria novamente
        ja_existe = (
            Notificacao.query
            .filter_by(usuario_id=usuario_id, link=link)
            .first()
        )
        if ja_existe:
            continue

        criar_notificacao(
            usuario_id=usuario_id,
            titulo="Agendamento para hoje",
            mensagem=f"Você tem um agendamento hoje às {hora_fmt} (Foco: {s.foco}).",
            link=link
        )


# -------------------------------------------------
# LER NOTIFICAÇÃO
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/ler")
@login_required
def ler_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    if n.lida_em is None:
        n.lida_em = agora_brasilia_naive()  # ✅ Brasília
        db.session.commit()

    return redirect(n.link or url_for("main.notificacoes"))


# -------------------------------------------------
# LISTAR NOTIFICAÇÕES
# -------------------------------------------------
@bp.route("/notificacoes")
@login_required
def notificacoes():
    user_tipo = current_user.tipo_usuario

    # ✅ só UVIS gera lembrete do dia (pro próprio usuário)
    if user_tipo not in ["admin", "operario", "visualizar"]:
        garantir_notificacoes_do_dia(current_user.id)

    base = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    # ✅ admin/operário/visualizar vê tudo, uvis só as dela
    if user_tipo in ["admin", "operario", "visualizar"]:
        itens = base.order_by(Notificacao.criada_em.desc()).all()
    else:
        itens = (base
                 .filter_by(usuario_id=current_user.id)
                 .order_by(Notificacao.criada_em.desc())
                 .all())

    return render_template("notificacoes.html", itens=itens)


# -------------------------------------------------
# EXCLUIR UMA NOTIFICAÇÃO (SOFT DELETE)
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/excluir", methods=["POST"])
@login_required
def excluir_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    n.apagada_em = agora_brasilia_naive()  # ✅ Brasília
    db.session.commit()

    return redirect(url_for("main.notificacoes"))


# -------------------------------------------------
# LIMPAR TODAS AS NOTIFICAÇÕES (SOFT DELETE EM LOTE)
# -------------------------------------------------
@bp.route("/notificacoes/limpar", methods=["POST"])
@login_required
def limpar_notificacoes():
    user_tipo = current_user.tipo_usuario
    agora = agora_brasilia_naive()  # ✅ Brasília

    q = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    if user_tipo not in ["admin", "operario", "visualizar"]:
        q = q.filter_by(usuario_id=current_user.id)

    q.update({"apagada_em": agora}, synchronize_session=False)
    db.session.commit()

    return redirect(url_for("main.notificacoes"))


# ==========================
# CHATBOT UVIS (FAQ inteligente)
# ==========================
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


UVIS_FAQ = [
    {
        "title": "Status da solicitação",
        "keywords": ["status", "pendente", "em analise", "aprovado", "negado", "protocolo"],
        "answer": (
            "📌 **Significado dos status**:\n"
            "- **Pendente**: solicitação registrada e aguardando início do processo.\n"
            "- **Em Análise**: pedido em validação pela equipe responsável.\n"
            "- **Aprovado**: pedido autorizado (pode aparecer o número de protocolo).\n"
            "- **Aprovado com Recomendações**: pedido aprovado com sugestões de melhoria.\n"
            "- **Negado**: pedido não aprovado (o motivo aparece nos detalhes).\n\n"
            "💡 Dica: clique em **Detalhes** para ver justificativa/protocolo."
        ),
    },
    {
        "title": "O que tem na tela 'Minhas Solicitações' (Dashboard)",
        "keywords": ["dashboard", "minhas solicitacoes", "tela inicial", "filtro", "detalhes", "nova solicitacao"],
        "answer": (
            "Na tela **Minhas Solicitações** você encontra:\n"
            "- Botão **Nova Solicitação** (abre o formulário)\n"
            "- **Filtro por status** (Pendente, Em Análise, Aprovado, Aprovado com Recomendações, Negado)\n"
            "- **Tabela** com data/hora, localização e foco\n"
            "- Botão **Detalhes** (abre um modal com informações completas)\n"
        ),
    },
    {
        "title": "Campos obrigatórios ao criar uma solicitação",
        "keywords": ["novo", "nova solicitacao", "cadastro", "campos", "obrigatorio", "cep", "numero", "tipo de visita", "altura", "foco"],
        "answer": (
            "✅ No cadastro de uma nova solicitação, atenção aos campos:\n"
            "- **Data** e **Hora** (obrigatórios)\n"
            "- **CEP** (8 dígitos) para preencher endereço automático\n"
            "- **Logradouro** (confirmar) e **Número** (preencher manualmente)\n"
            "- **Tipo de visita** (Monitoramento / Aedes / Culex)\n"
            "- **Altura do voo** (10m, 20m, 30m, 40m)\n"
            "- **Foco da ação** (ex.: Imóvel Abandonado, Piscina/Caixa d’água, Terreno Baldio, Ponto Estratégico)\n"
        ),
    },
    {
        "title": "CEP / endereço não encontrado e boas práticas",
        "keywords": ["cep", "endereco", "logradouro", "bairro", "cidade", "uf", "nao encontrado", "boas praticas"],
        "answer": (
            "Se o **CEP não for encontrado**, preencha o endereço manualmente e revise.\n"
            "Boas práticas:\n"
            "- confira se o **CEP** corresponde ao local\n"
            "- verifique logradouro/bairro/cidade/UF\n"
            "- preencha o **número** (sem ele pode dificultar a localização)\n"
        ),
    },
    {
        "title": "Latitude/Longitude e mapa",
        "keywords": ["latitude", "longitude", "coordenadas", "gps", "mapa"],
        "answer": (
            "📍 **Latitude/Longitude** é opcional (recomendado) e melhora a precisão.\n"
            "Se houver coordenadas, o sistema pode oferecer acesso rápido ao mapa."
        ),
    },
    {
        "title": "Notificações e Agenda",
        "keywords": ["notificacao", "notificacoes", "agenda", "calendario", "lembrete"],
        "answer": (
            "🔔 Em **Notificações**, você vê alertas da unidade (lembretes do dia/atualizações).\n"
            "Ao clicar, pode ser direcionado para a **Agenda**, que mostra os agendamentos por mês/semana/lista."
        ),
    },
    {
        "title": "Checklist antes de enviar",
        "keywords": ["checklist", "antes de enviar", "enviar pedido", "validar"],
        "answer": (
            "🧾 **Checklist rápido antes de enviar**:\n"
            "☐ Data e hora corretas\n"
            "☐ CEP válido e endereço conferido\n"
            "☐ Número preenchido\n"
            "☐ Tipo de visita e altura do voo selecionados\n"
            "☐ Foco da ação selecionado\n"
            "☐ Observações (se necessário) com informações objetivas\n"
        ),
    },
    {
        "title": "Suporte",
        "keywords": ["suporte", "erro", "acesso", "login", "senha"],
        "answer": (
            "Se a dúvida for de **erro de acesso**, inconsistência de **CEP/endereço**, ou algo fora do fluxo: "
            "entre em contato com o time de desenvolvimento/suporte da IJA."
        ),
    },
]


@bp.route("/api/uvis/chatbot", methods=["POST"])
@login_required
def uvis_chatbot():
    # (opcional) se quiser limitar só para UVIS:
    # if current_user.tipo_usuario != "uvis":
    #     return jsonify({"answer": "Acesso negado."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Escreva sua dúvida (ex.: “o que significa Em Análise?”)."}), 400

    nmsg = _norm(msg)

    best = None
    best_score = 0

    for item in UVIS_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "• “O que significa Pendente/Em Análise/Aprovado/Aprovado com Recomendações/Negado?”",
            "• “Quais campos são obrigatórios na Nova Solicitação?”",
            "• “O que fazer se o CEP não encontrar?”",
            "• “Qual o checklist antes de enviar?”",
            "• “Como funciona Notificações e Agenda?”",
        ]
        return jsonify({
            "answer": (
                "Não encontrei essa dúvida diretamente no manual.\n\n"
                "Tenta uma dessas perguntas:\n" + "\n".join(sugestoes)
            ),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": best["answer"],
        "matched": best["title"],
        "confidence": best_score,
    }), 200


import os
from flask import abort, send_from_directory
from flask_login import login_required, current_user

@bp.route("/solicitacao/<int:id>/anexo", endpoint="baixar_anexo")
@bp.route("/admin/solicitacao/<int:id>/anexo", endpoint="baixar_anexo_admin")
@login_required
def baixar_anexo(id):
    pedido = Solicitacao.query.get_or_404(id)

    # 🔐 permissões
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar", "uvis"]:
        abort(403)
    if current_user.tipo_usuario == "uvis" and pedido.usuario_id != current_user.id:
        abort(403)

    if not pedido.anexo_path:
        abort(404)

    # ✅ mesma pasta do upload
    upload_folder = get_upload_folder()

    # ✅ normaliza o caminho salvo no banco
    rel = (pedido.anexo_path or "").replace("\\", "/")
    if rel.startswith("upload-files/"):
        rel = rel.split("upload-files/", 1)[1]
    rel = os.path.basename(rel)  # segurança

    file_path = os.path.join(upload_folder, rel)
    if not os.path.isfile(file_path):
        abort(404)

    return send_from_directory(
        upload_folder,
        rel,
        as_attachment=False,
        download_name=(pedido.anexo_nome or rel)
    )

@bp.route("/admin/solicitacao/<int:id>/remover_anexo", methods=["POST"])
@login_required
def remover_anexo(id):
    pedido = Solicitacao.query.get_or_404(id)

    # ... lógica de permissão ...

    pedido.anexo_path = None
    pedido.anexo_nome = None
    db.session.commit()

    # ✅ Isso fará o Toast "Removido com sucesso" aparecer no topo igual aos outros deletes
    flash('PDF removido com sucesso!', 'success') 
    return redirect(url_for('main.dashboard'))

@bp.route("/admin/uvis/novo", methods=["GET", "POST"], endpoint="admin_uvis_novo")
@login_required
def admin_uvis_novo():
    # SOMENTE ADMIN
    if current_user.tipo_usuario != "admin":
        abort(403)

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        codigo_setor = (request.form.get("codigo_setor") or "").strip() or None

        login = (request.form.get("login") or "").strip()
        senha = request.form.get("senha") or ""
        confirmar = request.form.get("confirmar") or ""

        if not nome_uvis or not login or not senha:
            flash("Preencha: Nome da UVIS, Login e Senha.", "warning")
            return render_template("admin_uvis_novo.html")

        if senha != confirmar:
            flash("As senhas não conferem.", "warning")
            return render_template("admin_uvis_novo.html")

        novo_user = Usuario(
            nome_uvis=nome_uvis,
            regiao=regiao,
            codigo_setor=codigo_setor,
            login=login,
            tipo_usuario="uvis",
        )
        novo_user.set_senha(senha)

        try:
            db.session.add(novo_user)
            db.session.commit()
            flash("UVIS cadastrada com sucesso!", "success")
            return redirect(url_for("main.admin_uvis_listar"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar UVIS: {e}", "danger")

    return render_template("admin_uvis_novo.html")


@bp.route("/admin/uvis", methods=["GET"], endpoint="admin_uvis_listar")
@login_required
def admin_uvis_listar():
    # SOMENTE ADMIN
    if current_user.tipo_usuario != "admin":
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()
    codigo_setor = (request.args.get("codigo_setor") or "").strip()

    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")

    if q:
        query = query.filter(
            db.or_(
                Usuario.nome_uvis.ilike(f"%{q}%"),
                Usuario.login.ilike(f"%{q}%")
            )
        )

    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))

    if codigo_setor:
        query = query.filter(Usuario.codigo_setor.ilike(f"%{codigo_setor}%"))

    total = query.count()
    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(Usuario.nome_uvis.asc()).paginate(
        page=page, per_page=10, error_out=False
    )

    query = db.session.query(Solicitacao).options(
        joinedload(Solicitacao.usuario),
        joinedload(Solicitacao.piloto)  # ✅
    ).filter(Solicitacao.usuario_id == current_user.id)


    filters = {
        "q": q,
        "regiao": regiao,
        "codigo_setor": codigo_setor,
        "total": total
    }

    return render_template(
        "admin_uvis_listar.html",
        uvis=paginacao.items,
        paginacao=paginacao,
        filters=filters,
        q=q,
        regiao=regiao,
        codigo_setor=codigo_setor
    )

@bp.route("/admin/uvis/<int:id>/editar", methods=["GET", "POST"], endpoint="admin_uvis_editar")
@login_required
def admin_uvis_editar(id):
    if current_user.tipo_usuario != "admin":
        abort(403)

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para edição.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        codigo_setor = (request.form.get("codigo_setor") or "").strip() or None
        login = (request.form.get("login") or "").strip()

        senha = (request.form.get("senha") or "").strip()
        confirmar = (request.form.get("confirmar") or "").strip()

        if not nome_uvis or not login:
            flash("Preencha: Nome da UVIS e Login.", "warning")
            return render_template("admin_uvis_editar.html", uvis=uvis)

        if senha:
            if senha != confirmar:
                flash("As senhas não conferem.", "warning")
                return render_template("admin_uvis_editar.html", uvis=uvis)
            uvis.set_senha(senha)

        uvis.nome_uvis = nome_uvis
        uvis.regiao = regiao
        uvis.codigo_setor = codigo_setor
        uvis.login = login

        try:
            db.session.commit()
            flash("UVIS atualizada com sucesso!", "success")
            return redirect(url_for("main.admin_uvis_listar"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")

    return render_template("admin_uvis_editar.html", uvis=uvis)


@bp.route("/admin/uvis/<int:id>/excluir", methods=["POST"], endpoint="admin_uvis_excluir")
@login_required
def admin_uvis_excluir(id):
    if current_user.tipo_usuario != "admin":
        abort(403)

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para exclusão.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    existe = Solicitacao.query.filter_by(usuario_id=uvis.id).first()
    if existe:
        flash("Não é possível excluir: esta UVIS possui solicitações vinculadas.", "warning")
        return redirect(url_for("main.admin_uvis_listar"))

    try:
        db.session.delete(uvis)
        db.session.commit()
        flash("UVIS excluída com sucesso!", "success")
    except Exception:
        db.session.rollback()
        flash("Erro ao excluir UVIS.", "danger")

    return redirect(url_for("main.admin_uvis_listar"))

# ==========================
# CHATBOT ADMIN (FAQ inteligente) - Flask-Login
# ==========================
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm_admin(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_answer(text: str) -> str:
    """Remove markdown simples (**negrito**, `code`, etc) e normaliza."""
    if not text:
        return ""
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)   # remove ** **
    text = text.replace("`", "")                  # remove ` `
    text = re.sub(r"\n{3,}", "\n\n", text)         # evita muitas quebras
    return text.strip()


ADMIN_FAQ = [
    {
        "title": "Perfis e permissões",
        "keywords": ["acesso", "perfil", "permissao", "permissões", "admin", "operario", "operário", "visualizar", "quem pode"],
        "answer": (
            "Perfis do painel:\n"
            "- Administrador: acesso total (editar, excluir, gerenciar UVIS, relatórios e agenda).\n"
            "- Operário: consegue salvar decisões (status/protocolo/justificativa).\n"
            "- Visualizar: apenas leitura.\n"
        ),
    },
    {
        "title": "Filtros no painel",
        "keywords": ["filtro", "filtrar", "status", "unidade", "uvis", "regiao", "região", "buscar", "pesquisar"],
        "answer": (
            "No painel você pode filtrar por:\n"
            "- Status\n"
            "- Unidade (UVIS)\n"
            "- Região\n"
            "Use os filtros para encontrar solicitações específicas rapidamente."),
    },
    {
        "title": "Olá! Como posso ajudar?",
        "keywords": ["olá", "oi", "hello", "hi", "bom dia", "boa tarde", "boa noite", "ajuda", "suporte"],
        "answer": (
            "Olá! Sou o assistente virtual do painel administrativo.\n"
            "Posso ajudar com dúvidas sobre:\n"
            "- Perfis e permissões\n"
            "- Filtros no painel\n"
            "- Salvar decisão\n"
            "- Editar completo\n"
            "- Excluir solicitação\n"
            "- Anexos\n"
            "- GPS e mapa\n"
            "- Exportar Excel do painel\n"
            "- Agenda\n"
            "- Relatórios\n"
            "- Gestão de UVIS\n"
            "Como posso ajudar você hoje?"
        ),
    },
    {
        "title": "Salvar decisão",
        "keywords": ["salvar", "decisao", "decisão", "status", "protocolo", "justificativa", "aprovado", "negado", "analise", "recomendacoes", "recomendações"],
        "answer": (
            "Em cada solicitação você pode definir:\n"
            "- Status\n"
            "- Protocolo\n"
            "- Justificativa (principalmente se negar ou orientar)\n"
            "Se o perfil for ‘Visualizar’, fica somente leitura."
        ),
    },
    {
        "title": "Editar completo",
        "keywords": ["editar", "editar completo", "corrigir", "alterar", "data", "hora", "endereco", "endereço", "agendamento"],
        "answer": (
            "Editar completo serve para corrigir todos os dados do pedido:\n"
            "data/hora, endereço, foco, tipo de visita, altura e observações.\n"
            "Em alguns casos o sistema pode gerar notificação para a unidade."
        ),
    },
    {
        "title": "Excluir solicitação",
        "keywords": ["excluir", "deletar", "apagar", "remover"],
        "answer": (
            "Excluir remove a solicitação definitivamente.\n"
            "Normalmente é restrito ao Administrador e pede confirmação."
        ),
    },
    {
        "title": "Anexos",
        "keywords": ["anexo", "arquivo", "upload", "baixar", "download", "pdf", "png", "jpg", "doc", "xlsx"],
        "answer": (
            "Você pode anexar arquivos na solicitação e depois baixar.\n"
            "Se o anexo não aparecer, verifique se foi salvo corretamente e se o arquivo é permitido."
        ),
    },
    {
        "title": "GPS e mapa",
        "keywords": ["gps", "latitude", "longitude", "coordenadas", "mapa", "google maps"],
        "answer": (
            "Latitude/Longitude ajudam na precisão.\n"
            "Quando preenchidas, o botão de mapa abre o local no Google Maps."
        ),
    },
    {
        "title": "Exportar Excel do painel",
        "keywords": ["exportar", "excel", "xlsx", "planilha", "baixar excel"],
        "answer": (
            "Existe exportação para Excel a partir do painel.\n"
            "Quando você usa filtros (status/unidade/região), isso tende a refletir no arquivo exportado."
        ),
    },
    {
        "title": "Agenda",
        "keywords": ["agenda", "calendario", "calendário", "eventos", "mes", "mês", "ano", "exportar agenda"],
        "answer": (
            "A Agenda mostra agendamentos por período.\n"
            "Você pode filtrar (quando disponível) e exportar."
        ),
    },
    {
        "title": "Relatórios",
        "keywords": ["relatorio", "relatórios", "pdf", "grafico", "gráfico", "totais", "mes", "ano"],
        "answer": (
            "Relatórios permitem filtrar por mês/ano e, quando disponível, por unidade.\n"
            "Também podem ter exportação em PDF e Excel."
        ),
    },
    {
        "title": "Gestão de UVIS",
        "keywords": ["uvis", "cadastrar uvis", "lista uvis", "gerenciar uvis", "unidade", "login", "senha", "codigo setor", "código setor", "regiao", "região"],
        "answer": (
            "Gestão de UVIS inclui:\n"
            "- Listar UVIS\n"
            "- Cadastrar UVIS\n"
            "- Editar UVIS (inclusive redefinir senha)\n"
            "Atenção: login não pode repetir."
        ),
    },
]


@bp.route("/api/admin/chatbot", methods=["POST"])
@login_required
def admin_chatbot():
    # 🔐 só perfis do painel
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
        return jsonify({"answer": "Acesso negado para este chatbot."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Digite sua dúvida (ex.: como exportar Excel?)."}), 400

    nmsg = _norm_admin(msg)

    best = None
    best_score = 0

    for item in ADMIN_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "Como filtrar por status/unidade/região?",
            "Como salvar decisão (status/protocolo/justificativa)?",
            "Como editar completo?",
            "Como exportar Excel?",
            "Como funciona Agenda/Relatórios?",
            "Como gerenciar UVIS?",
        ]
        return jsonify({
            "answer": "Não achei essa dúvida direto no guia.\n\nSugestões:\n- " + "\n- ".join(sugestoes),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": _clean_answer(best["answer"]),
        "matched": best["title"],
        "confidence": best_score,
    }), 200



@bp.app_errorhandler(404)
def pagina_nao_encontrada(e):
    return render_template(
        'erro.html', 
        codigo=404, 
        titulo="Página não encontrada", 
        mensagem="Ops! A página que você está procurando não existe ou foi movida."
    ), 404

@bp.app_errorhandler(500)
def erro_interno(e):
    # Opcional: printar o erro no terminal para você ver o que houve
    # print(f"Erro 500 detectado: {e}")
    return render_template(
        'erro.html', 
        codigo=500, 
        titulo="Erro Interno do Servidor", 
        mensagem="Desculpe, algo deu errado do nosso lado. Tente novamente mais tarde."
    ), 500

import re
import requests
from flask import jsonify, current_app
from flask_login import login_required

def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")

@bp.route("/api/cep/<cep>", methods=["GET"], endpoint="api_cep")
@login_required
def api_cep(cep):
    cep_digits = only_digits(cep)

    if len(cep_digits) != 8:
        return jsonify(ok=False, error="CEP inválido. Use 8 dígitos."), 400

    def _resp_ok(payload):
        return jsonify(
            ok=True,
            cep=payload.get("cep", ""),
            logradouro=payload.get("logradouro", ""),
            complemento=payload.get("complemento", ""),
            bairro=payload.get("bairro", ""),
            cidade=payload.get("cidade", ""),
            uf=payload.get("uf", ""),
        )

    headers = {"User-Agent": "Mozilla/5.0"}

    # 1) ViaCEP
    try:
        r = requests.get(f"https://viacep.com.br/ws/{cep_digits}/json/", timeout=8, headers=headers)
        r.raise_for_status()
        data = r.json()

        if data.get("erro"):
            return jsonify(ok=False, error="CEP não encontrado."), 404

        payload = {
            "cep": data.get("cep", ""),
            "logradouro": data.get("logradouro", ""),
            "complemento": data.get("complemento", ""),
            "bairro": data.get("bairro", ""),
            "cidade": data.get("localidade", ""),
            "uf": data.get("uf", ""),
        }
        return _resp_ok(payload)

    except Exception as e:
        current_app.logger.exception("Falha ViaCEP: %s", e)

        # 2) Fallback: BrasilAPI
        try:
            r2 = requests.get(f"https://brasilapi.com.br/api/cep/v1/{cep_digits}", timeout=8, headers=headers)
            r2.raise_for_status()
            data2 = r2.json()

            payload = {
                "cep": data2.get("cep", ""),
                "logradouro": data2.get("street", ""),
                "complemento": "",  # BrasilAPI normalmente não traz
                "bairro": data2.get("neighborhood", ""),
                "cidade": data2.get("city", ""),
                "uf": data2.get("state", ""),
            }
            return _resp_ok(payload)

        except Exception as e2:
            current_app.logger.exception("Falha BrasilAPI: %s", e2)

            # Se estiver em DEBUG, mostra o erro real pra você ver a causa
            if current_app.debug:
                return jsonify(ok=False, error=f"Falha CEP (debug): {repr(e2)}"), 502

            return jsonify(ok=False, error="Falha ao consultar o serviço de CEP."), 502

# -----------------------------
# Helpers: CPF/CNPJ/Telefone/CEP
# -----------------------------
def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")

def format_cpf(cpf_digits: str) -> str:
    return f"{cpf_digits[:3]}.{cpf_digits[3:6]}.{cpf_digits[6:9]}-{cpf_digits[9:11]}"

def format_cnpj(cnpj_digits: str) -> str:
    return f"{cnpj_digits[:2]}.{cnpj_digits[2:5]}.{cnpj_digits[5:8]}/{cnpj_digits[8:12]}-{cnpj_digits[12:14]}"

def validate_cpf(cpf: str) -> bool:
    cpf = only_digits(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    d1 = (soma * 10) % 11
    d1 = 0 if d1 == 10 else d1
    if d1 != int(cpf[9]):
        return False

    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    d2 = (soma * 10) % 11
    d2 = 0 if d2 == 10 else d2
    return d2 == int(cpf[10])

def validate_cnpj(cnpj: str) -> bool:
    cnpj = only_digits(cnpj)
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    pesos2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]

    soma = sum(int(cnpj[i]) * pesos1[i] for i in range(12))
    d1 = 11 - (soma % 11)
    d1 = 0 if d1 >= 10 else d1
    if d1 != int(cnpj[12]):
        return False

    soma = sum(int(cnpj[i]) * pesos2[i] for i in range(13))
    d2 = 11 - (soma % 11)
    d2 = 0 if d2 >= 10 else d2
    return d2 == int(cnpj[13])

def validate_documento(doc: str):
    """
    Retorna (ok, tipo, doc_digits, doc_formatado, erro_msg)
    tipo: 'CPF' | 'CNPJ'
    """
    digits = only_digits(doc)

    if len(digits) == 11:
        if not validate_cpf(digits):
            return False, "CPF", digits, None, "CPF inválido (dígitos verificadores não conferem)."
        return True, "CPF", digits, format_cpf(digits), None

    if len(digits) == 14:
        if not validate_cnpj(digits):
            return False, "CNPJ", digits, None, "CNPJ inválido (dígitos verificadores não conferem)."
        return True, "CNPJ", digits, format_cnpj(digits), None

    return False, None, digits, None, "Documento deve ter 11 (CPF) ou 14 (CNPJ) dígitos."

def format_phone_br(phone_digits: str) -> str:
    d = only_digits(phone_digits)
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:11]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:10]}"
    return phone_digits

def format_cep(cep_digits: str) -> str:
    d = only_digits(cep_digits)
    if len(d) == 8:
        return f"{d[:5]}-{d[5:]}"
    return cep_digits

def build_endereco_full(cep, logradouro, numero, complemento, bairro, cidade, uf) -> str:
    cep = only_digits(cep)
    logradouro = (logradouro or "").strip()
    numero = (numero or "").strip()
    complemento = (complemento or "").strip()
    bairro = (bairro or "").strip()
    cidade = (cidade or "").strip()
    uf = (uf or "").strip().upper()

    linha1 = ""
    if logradouro:
        linha1 += logradouro
    if numero:
        linha1 += f", {numero}" if linha1 else numero
    if complemento:
        linha1 += f" ({complemento})" if linha1 else complemento

    cidade_uf = ""
    if cidade and uf:
        cidade_uf = f"{cidade}/{uf}"
    else:
        cidade_uf = cidade or uf

    linha2 = " - ".join([x for x in [bairro, cidade_uf] if x])

    cep_fmt = format_cep(cep) if cep else ""
    linha3 = f"CEP {cep_fmt}" if cep_fmt else ""

    return " - ".join([x for x in [linha1, linha2, linha3] if x]).strip()


# -----------------------------
# Rota: cadastrar clientes (COM CEP)
# -----------------------------
@bp.route('/clientes/cadastrar', methods=['GET', 'POST'], endpoint='cadastrar_clientes')
@login_required
def cadastrar_clientes():
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    errors = {}
    form = {}

    if request.method == "POST":
        nome_cliente = (request.form.get("nome_cliente") or "").strip()
        documento = (request.form.get("documento") or "").strip()
        contato = (request.form.get("contato") or "").strip()
        telefone = (request.form.get("telefone") or "").strip()
        email = (request.form.get("email") or "").strip()

        # NOVOS CAMPOS DE ENDEREÇO
        cep = (request.form.get("cep") or "").strip()
        logradouro = (request.form.get("logradouro") or "").strip()
        numero = (request.form.get("numero") or "").strip()
        complemento = (request.form.get("complemento") or "").strip()
        bairro = (request.form.get("bairro") or "").strip()
        cidade = (request.form.get("cidade") or "").strip()
        uf = (request.form.get("uf") or "").strip().upper()

        # Fallback (caso ainda exista campo antigo no form)
        endereco_raw = (request.form.get("endereco") or "").strip()

        # Monta o endereço final (se qualquer campo novo veio preenchido)
        tem_endereco_novo = any([cep, logradouro, numero, complemento, bairro, cidade, uf])
        endereco_full = build_endereco_full(cep, logradouro, numero, complemento, bairro, cidade, uf) if tem_endereco_novo else endereco_raw

        # Mantém valores pra re-render do form
        form = {
            "nome_cliente": nome_cliente,
            "documento": documento,
            "contato": contato,
            "telefone": telefone,
            "email": email,

            # novos
            "cep": cep,
            "logradouro": logradouro,
            "numero": numero,
            "complemento": complemento,
            "bairro": bairro,
            "cidade": cidade,
            "uf": uf,

            # compat
            "endereco": endereco_full,
        }

        # Obrigatórios
        if not nome_cliente:
            errors["nome_cliente"] = "Informe o nome do cliente."
        if not documento:
            errors["documento"] = "Informe CPF ou CNPJ."

        # Documento (CPF/CNPJ)
        doc_ok, doc_tipo, doc_digits, doc_fmt, doc_err = validate_documento(documento)
        if documento and not doc_ok:
            errors["documento"] = doc_err

        # Email (se preenchido)
        if email:
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
                errors["email"] = "E-mail inválido. Ex: nome@dominio.com"

        # Telefone (se preenchido)
        if telefone:
            tel_digits = only_digits(telefone)
            if len(tel_digits) not in (10, 11):
                errors["telefone"] = "Telefone deve ter 10 ou 11 dígitos (com DDD)."

        # CEP (se preenchido)
        if cep:
            cep_digits = only_digits(cep)
            if len(cep_digits) != 8:
                errors["cep"] = "CEP deve ter 8 dígitos."

        # Se doc válido, checar duplicidade antes de tentar salvar
        if doc_ok:
            existe = Clientes.query.filter_by(documento=doc_digits).first()
            if existe:
                errors["documento"] = f"Já existe um cliente cadastrado com esse {doc_tipo}."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("cadastrar_clientes.html", form=form, errors=errors)

        # Salvar (documento e telefone SEM máscara)
        novo = Clientes(
            nome_cliente=nome_cliente,
            documento=doc_digits,
            contato=contato or None,
            telefone=only_digits(telefone) or None,
            email=email or None,
            endereco=endereco_full or None
        )

        db.session.add(novo)
        db.session.commit()

        flash(f"Cliente cadastrado com sucesso! Documento salvo como {doc_fmt}.", "success")
        return redirect(url_for("main.listar_clientes"))

    return render_template("cadastrar_clientes.html", form=form, errors=errors)


import math
import re
from io import BytesIO
from datetime import datetime

from flask import render_template, abort, request, send_file
from flask_login import login_required, current_user

from app import db
from app.models import Clientes


# -----------------------------
# Helpers: CPF/CNPJ/Telefone
# -----------------------------
def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def format_cpf(cpf_digits: str) -> str:
    d = only_digits(cpf_digits)
    if len(d) != 11:
        return cpf_digits
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"


def format_cnpj(cnpj_digits: str) -> str:
    d = only_digits(cnpj_digits)
    if len(d) != 14:
        return cnpj_digits
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def format_documento(doc: str) -> str:
    d = only_digits(doc)
    if len(d) == 11:
        return format_cpf(d)
    if len(d) == 14:
        return format_cnpj(d)
    return doc


def format_phone_br(phone: str) -> str:
    d = only_digits(phone)
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:11]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:10]}"
    return phone


@bp.route("/clientes", methods=["GET"], endpoint="listar_clientes")
@login_required
def listar_clientes():
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    # -----------------------------
    # Params (filtros / paginação)
    # -----------------------------
    q = (request.args.get("q") or "").strip()
    doc = (request.args.get("doc") or "").strip()  # cpf/cnpj (com ou sem máscara)
    email = (request.args.get("email") or "").strip()
    telefone = (request.args.get("telefone") or "").strip()
    sort = (request.args.get("sort") or "nome_asc").strip()

    try:
        page = int(request.args.get("page") or 1)
    except ValueError:
        page = 1
    page = max(1, page)

    try:
        per_page = int(request.args.get("per_page") or 20)
    except ValueError:
        per_page = 20
    per_page = 10 if per_page < 10 else 50 if per_page > 50 else per_page

    export = (request.args.get("export") or "").strip().lower()  # "xlsx"

    # -----------------------------
    # Query base
    # -----------------------------
    query = Clientes.query

    # filtro documento (salvo como dígitos)
    if doc:
        query = query.filter(Clientes.documento.ilike(f"%{only_digits(doc)}%"))

    # filtro email
    if email:
        query = query.filter(Clientes.email.ilike(f"%{email}%"))

    # filtro telefone (salvo como dígitos)
    if telefone:
        query = query.filter(Clientes.telefone.ilike(f"%{only_digits(telefone)}%"))

    # busca geral (nome, contato, email, endereço, doc, telefone)
    if q:
        like = f"%{q}%"
        q_digits = only_digits(q)

        query = query.filter(
            db.or_(
                Clientes.nome_cliente.ilike(like),
                Clientes.contato.ilike(like),
                Clientes.email.ilike(like),
                Clientes.endereco.ilike(like),
                Clientes.documento.ilike(f"%{q_digits}%") if q_digits else db.false(),
                Clientes.telefone.ilike(f"%{q_digits}%") if q_digits else db.false(),
            )
        )

    # ordenação
    if sort == "nome_desc":
        query = query.order_by(Clientes.nome_cliente.desc())
    elif sort == "id_desc":
        query = query.order_by(Clientes.id.desc())
    elif sort == "id_asc":
        query = query.order_by(Clientes.id.asc())
    else:
        query = query.order_by(Clientes.nome_cliente.asc())

    # -----------------------------
    # Exportação Excel (filtrado)
    # -----------------------------
    if export == "xlsx":
        rows = query.all()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # --- Estilos ---
        header_fill = PatternFill("solid", fgColor="1F2937")  # cinza escuro
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # --- (Opcional) Título do relatório ---
        ws["A1"] = "Relatório de Clientes"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(color="6B7280")

        # Linha de cabeçalho começa na 4
        start_row = 4
        headers = ["ID", "Nome", "Documento", "Contato", "Telefone", "E-mail", "Endereço"]

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # --- Dados ---
        for i, c in enumerate(rows, start=start_row + 1):
            values = [
                c.id,
                c.nome_cliente,
                format_documento(c.documento),      # já formatado
                c.contato or "",
                format_phone_br(c.telefone or ""),  # já formatado
                c.email or "",
                c.endereco or "",
            ]

            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx, value=v)
                cell.border = border

                # alinhamento por coluna
                if col_idx == 1:  # ID
                    cell.alignment = center_align
                else:
                    cell.alignment = text_align

                # força texto em Documento/Telefone pra não virar número
                if col_idx in (3, 5):
                    cell.number_format = "@"

        last_row = start_row + len(rows)
        last_col = len(headers)

        # --- Congela cabeçalho ---
        ws.freeze_panes = ws["A5"]  # congela acima da linha 5

        # --- AutoFilter ---
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{max(last_row, start_row)}"

        # --- Altura do cabeçalho ---
        ws.row_dimensions[start_row].height = 22

        # --- Largura de colunas (mais inteligente) ---
        max_widths = {1: 8, 2: 28, 3: 22, 4: 18, 5: 18, 6: 26, 7: 45}

        for col_idx in range(1, last_col + 1):
            max_len = len(headers[col_idx - 1])
            for r in range(start_row + 1, last_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))

            width = min(max_len + 2, max_widths.get(col_idx, 40))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # --- Zebra (linhas alternadas) ---
        zebra_fill = PatternFill("solid", fgColor="F9FAFB")
        for r in range(start_row + 1, last_row + 1):
            if (r - (start_row + 1)) % 2 == 1:
                for ccol in range(1, last_col + 1):
                    ws.cell(row=r, column=ccol).fill = zebra_fill

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"clientes_{stamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -----------------------------
    # Paginação
    # -----------------------------
    total = query.count()
    total_pages = max(1, math.ceil(total / per_page))

    if page > total_pages:
        page = total_pages

    clientes_db = query.offset((page - 1) * per_page).limit(per_page).all()

    # manda já formatado pro template
    clientes = [
        {
            "id": c.id,
            "nome_cliente": c.nome_cliente,
            "documento_fmt": format_documento(c.documento),
            "contato": c.contato or "-",
            "telefone_fmt": format_phone_br(c.telefone or "") or "-",
            "email": c.email or "-",
            "endereco": c.endereco or "-",
        }
        for c in clientes_db
    ]

    filters = {
        "q": q,
        "doc": doc,
        "email": email,
        "telefone": telefone,
        "sort": sort,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
    }

    return render_template("listar_clientes.html", clientes=clientes, filters=filters)
from flask import render_template, request, redirect, url_for, flash, abort
from flask_login import login_required, current_user
from app import db
from app.models import Clientes
import re

# -----------------------------
# Helpers: digits / format
# -----------------------------
def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")

def format_cpf(cpf_digits: str) -> str:
    d = only_digits(cpf_digits)
    if len(d) != 11:
        return cpf_digits
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"

def format_cnpj(cnpj_digits: str) -> str:
    d = only_digits(cnpj_digits)
    if len(d) != 14:
        return cnpj_digits
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"

def format_documento(doc: str) -> str:
    d = only_digits(doc)
    if len(d) == 11:
        return format_cpf(d)
    if len(d) == 14:
        return format_cnpj(d)
    return doc

def format_phone_br(phone: str) -> str:
    d = only_digits(phone)
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:11]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:10]}"
    return phone


# -----------------------------
# Validação CPF/CNPJ
# -----------------------------
def validate_cpf(cpf: str) -> bool:
    cpf = only_digits(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    d1 = (soma * 10) % 11
    d1 = 0 if d1 == 10 else d1
    if d1 != int(cpf[9]):
        return False

    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    d2 = (soma * 10) % 11
    d2 = 0 if d2 == 10 else d2
    return d2 == int(cpf[10])

def validate_cnpj(cnpj: str) -> bool:
    cnpj = only_digits(cnpj)
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    pesos1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    pesos2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]

    soma = sum(int(cnpj[i]) * pesos1[i] for i in range(12))
    d1 = 11 - (soma % 11)
    d1 = 0 if d1 >= 10 else d1
    if d1 != int(cnpj[12]):
        return False

    soma = sum(int(cnpj[i]) * pesos2[i] for i in range(13))
    d2 = 11 - (soma % 11)
    d2 = 0 if d2 >= 10 else d2
    return d2 == int(cnpj[13])

def validate_documento(doc: str):
    """
    Retorna (ok, tipo, doc_digits, doc_formatado, erro_msg)
    tipo: 'CPF' | 'CNPJ'
    """
    digits = only_digits(doc)

    if len(digits) == 11:
        if not validate_cpf(digits):
            return False, "CPF", digits, None, "CPF inválido (dígitos verificadores não conferem)."
        return True, "CPF", digits, format_cpf(digits), None

    if len(digits) == 14:
        if not validate_cnpj(digits):
            return False, "CNPJ", digits, None, "CNPJ inválido (dígitos verificadores não conferem)."
        return True, "CNPJ", digits, format_cnpj(digits), None

    return False, None, digits, None, "Documento deve ter 11 (CPF) ou 14 (CNPJ) dígitos."


# -----------------------------
# EDITAR CLIENTE (admin)
# -----------------------------
@bp.route("/clientes/<int:cliente_id>/editar", methods=["GET", "POST"], endpoint="editar_cliente")
@login_required
def editar_cliente(cliente_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    cliente = Clientes.query.get_or_404(cliente_id)

    errors = {}
    form = {}

    if request.method == "POST":
        nome_cliente = (request.form.get("nome_cliente") or "").strip()
        documento = (request.form.get("documento") or "").strip()
        contato = (request.form.get("contato") or "").strip()
        telefone = (request.form.get("telefone") or "").strip()
        email = (request.form.get("email") or "").strip()
        endereco = (request.form.get("endereco") or "").strip()

        form = {
            "nome_cliente": nome_cliente,
            "documento": documento,
            "contato": contato,
            "telefone": telefone,
            "email": email,
            "endereco": endereco,
        }

        # obrigatórios
        if not nome_cliente:
            errors["nome_cliente"] = "Informe o nome do cliente."
        if not documento:
            errors["documento"] = "Informe CPF ou CNPJ."

        # documento
        doc_ok, doc_tipo, doc_digits, doc_fmt, doc_err = validate_documento(documento)
        if documento and not doc_ok:
            errors["documento"] = doc_err

        # email
        if email:
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
                errors["email"] = "E-mail inválido. Ex: nome@dominio.com"

        # telefone
        if telefone:
            tel_digits = only_digits(telefone)
            if len(tel_digits) not in (10, 11):
                errors["telefone"] = "Telefone deve ter 10 ou 11 dígitos (com DDD)."

        # duplicidade documento (ignora o próprio)
        if doc_ok:
            existe = (
                Clientes.query
                .filter(Clientes.documento == doc_digits, Clientes.id != cliente.id)
                .first()
            )
            if existe:
                errors["documento"] = f"Já existe outro cliente com esse {doc_tipo}."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("editar_cliente.html", form=form, errors=errors, cliente=cliente)

        # salva
        cliente.nome_cliente = nome_cliente
        cliente.documento = doc_digits
        cliente.contato = contato or None
        cliente.telefone = only_digits(telefone) or None
        cliente.email = email or None
        cliente.endereco = endereco or None

        db.session.commit()

        flash(f"Cliente atualizado! Documento: {doc_fmt}", "success")
        return redirect(url_for("main.listar_clientes"))

    # GET: preenche form com dados atuais (formatados)
    form = {
        "nome_cliente": cliente.nome_cliente,
        "documento": format_documento(cliente.documento),
        "contato": cliente.contato or "",
        "telefone": format_phone_br(cliente.telefone or ""),
        "email": cliente.email or "",
        "endereco": cliente.endereco or "",
    }

    return render_template("editar_cliente.html", form=form, errors=errors, cliente=cliente)


# -----------------------------
# DELETAR CLIENTE (admin)
# -----------------------------
@bp.route("/clientes/<int:cliente_id>/deletar", methods=["POST"], endpoint="deletar_cliente")
@login_required
def deletar_cliente(cliente_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    cliente = Clientes.query.get_or_404(cliente_id)

    db.session.delete(cliente)
    db.session.commit()

    flash("Cliente removido com sucesso.", "success")
    return redirect(url_for("main.listar_clientes"))
from flask import request, abort, send_file
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@bp.route("/admin/uvis/exportar", methods=["GET"], endpoint="admin_uvis_exportar")
@login_required
def admin_uvis_exportar():
    if current_user.tipo_usuario != "admin":
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()
    codigo_setor = (request.args.get("codigo_setor") or "").strip()

    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")

    if q:
        query = query.filter(
            db.or_(
                Usuario.nome_uvis.ilike(f"%{q}%"),
                Usuario.login.ilike(f"%{q}%")
            )
        )
    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))
    if codigo_setor:
        query = query.filter(Usuario.codigo_setor.ilike(f"%{codigo_setor}%"))

    rows = query.order_by(Usuario.nome_uvis.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "UVIS"

    # ---------- ESTILOS ----------
    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10, color="666666")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    zebra_fill = PatternFill("solid", fgColor="F3F6FA")

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # ---------- TÍTULO / META (FORA DA TABELA) ----------
    ws["A1"] = "UVIS Cadastradas"
    ws["A1"].font = title_font

    ws["A3"] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = meta_font

    start_header_row = 5

    # ---------- CABEÇALHO ----------
    headers = ["ID", "Nome", "Região", "Cód. Setor", "Login"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # ---------- DADOS ----------
    start_data_row = start_header_row + 1
    for i, u in enumerate(rows):
        r = start_data_row + i
        values = [u.id, u.nome_uvis, u.regiao, u.codigo_setor, u.login]

        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center if c == 1 else left

            if i % 2 == 1:
                cell.fill = zebra_fill

    end_data_row = start_data_row + len(rows) - 1

    # ---------- AUTOFILTER (SEGURO) ----------
    if rows:
        ws.auto_filter.ref = f"A{start_header_row}:E{end_data_row}"
        ws.freeze_panes = f"A{start_data_row}"

    # ---------- LARGURAS ----------
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24

    # ---------- TOTAL ----------
    total_row = end_data_row + 2
    ws.cell(row=total_row, column=1, value="Total de UVIS:").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(rows)).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"uvis_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route('/sw.js')
def serve_sw():
    return bp.send_static_file('sw.js')
# se você já tem essa função no projeto, reaproveite
def only_digits(v: str) -> str:
    return re.sub(r"\D+", "", v or "")

import re
import math
from datetime import datetime
from io import BytesIO

from flask import request, render_template, flash, redirect, url_for, abort, send_file
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash

from app import db
from app.models import Pilotos, Usuario  # ajuste o import conforme sua estrutura


def only_digits(v: str) -> str:
    return re.sub(r"\D+", "", v or "")


def format_phone_br(digits: str) -> str:
    d = only_digits(digits)
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    return digits or ""


REGIOES = {"NORTE", "SUL", "LESTE", "OESTE"}


# -------------------------------------------------------------
# CADASTRAR PILOTO + CRIAR USUÁRIO (tipo_usuario="piloto")
# -------------------------------------------------------------
@bp.route('/pilotos/cadastrar', methods=['GET', 'POST'], endpoint='cadastrar_pilotos')
@login_required
def cadastrar_pilotos():
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    errors = {}
    form = {}

    if request.method == "POST":
        # dados do piloto
        nome_piloto = (request.form.get("nome_piloto") or "").strip()
        regiao = (request.form.get("regiao") or "").strip().upper()
        telefone = (request.form.get("telefone") or "").strip()
        tel_digits = only_digits(telefone)

        # credenciais do usuário piloto
        login = (request.form.get("login") or "").strip()
        senha = (request.form.get("senha") or "")
        senha2 = (request.form.get("senha2") or "")

        form = {
            "nome_piloto": nome_piloto,
            "regiao": regiao,
            "telefone": telefone,
            "login": login,
            "senha": senha,
            "senha2": senha2,
        }

        # validações
        if not nome_piloto:
            errors["nome_piloto"] = "Informe o nome do piloto."

        if regiao and regiao not in REGIOES:
            errors["regiao"] = "Selecione uma região válida."

        if telefone and len(tel_digits) not in (10, 11):
            errors["telefone"] = "Telefone deve ter 10 ou 11 dígitos (com DDD)."

        # duplicidade de piloto (nome + telefone se tiver)
        if nome_piloto:
            q = Pilotos.query.filter(db.func.lower(Pilotos.nome_piloto) == nome_piloto.lower())
            if tel_digits:
                q = q.filter(Pilotos.telefone == tel_digits)
            if q.first():
                errors["nome_piloto"] = "Já existe um piloto com esse nome (e telefone)."

        # login
        if not login:
            errors["login"] = "Informe um login para o piloto."
        else:
            existe_login = Usuario.query.filter(db.func.lower(Usuario.login) == login.lower()).first()
            if existe_login:
                errors["login"] = "Esse login já está em uso."

        # senha
        if not senha:
            errors["senha"] = "Informe uma senha."
        elif len(senha) < 6:
            errors["senha"] = "A senha deve ter pelo menos 6 caracteres."

        if senha != senha2:
            errors["senha2"] = "As senhas não conferem."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template("cadastrar_pilotos.html", form=form, errors=errors)

        try:
            # 1) cria piloto
            novo_piloto = Pilotos(
                nome_piloto=nome_piloto,
                regiao=regiao or None,
                telefone=tel_digits or None,
            )
            db.session.add(novo_piloto)
            db.session.flush()  # garante novo_piloto.id

            # 2) cria usuário do piloto (nome_uvis é obrigatório no seu model)
            user_piloto = Usuario(
                nome_uvis=nome_piloto,          # ✅ obrigatório
                regiao=regiao or None,          # opcional, mas útil
                codigo_setor=None,
                login=login,
                tipo_usuario="piloto",
                piloto_id=novo_piloto.id,
            )
            user_piloto.set_senha(senha)

            db.session.add(user_piloto)
            db.session.commit()

            flash("Piloto e usuário criados com sucesso!", "success")
            return redirect(url_for("main.listar_pilotos"))

        except Exception:
            db.session.rollback()
            flash("Erro ao cadastrar piloto/usuário. Tente novamente.", "danger")
            return render_template("cadastrar_pilotos.html", form=form, errors=errors)

    return render_template("cadastrar_pilotos.html", form=form, errors=errors)


# -------------------------------------------------------------
# LISTAR PILOTOS (admin e uvis)
# -------------------------------------------------------------
@bp.route("/pilotos", methods=["GET"], endpoint="listar_pilotos")
@login_required
def listar_pilotos():
    user_tipo = getattr(current_user, "tipo_usuario", None)

    if user_tipo not in ("admin", "uvis"):
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip().upper()
    telefone = (request.args.get("telefone") or "").strip()
    sort = (request.args.get("sort") or "nome_asc").strip()

    try:
        page = int(request.args.get("page") or 1)
    except ValueError:
        page = 1
    page = max(1, page)

    try:
        per_page = int(request.args.get("per_page") or 20)
    except ValueError:
        per_page = 20
    per_page = 10 if per_page < 10 else 50 if per_page > 50 else per_page

    export = (request.args.get("export") or "").strip().lower()

    # ✅ Controle de região para UVIS (força a regiao da uvis)
    uvis_regiao = (getattr(current_user, "regiao", None) or "").strip().upper()
    if user_tipo == "uvis":
        if not uvis_regiao:
            flash("Sua UVIS está sem região cadastrada. Contate o administrador.", "warning")
            return render_template(
                "listar_pilotos.html",
                pilotos=[],
                filters={"q": q, "regiao": "", "telefone": telefone, "sort": sort, "page": 1, "per_page": per_page, "total": 0, "total_pages": 1},
                is_admin=False,
                uvis_regiao=None
            )
        regiao = uvis_regiao

    query = Pilotos.query

    if regiao:
        query = query.filter(Pilotos.regiao == regiao)

    if telefone:
        query = query.filter(Pilotos.telefone.ilike(f"%{only_digits(telefone)}%"))

    if q:
        like = f"%{q}%"
        q_digits = only_digits(q)
        query = query.filter(
            db.or_(
                Pilotos.nome_piloto.ilike(like),
                Pilotos.regiao.ilike(like),
                Pilotos.telefone.ilike(f"%{q_digits}%") if q_digits else db.false(),
            )
        )

    if sort == "nome_desc":
        query = query.order_by(Pilotos.nome_piloto.desc())
    elif sort == "id_desc":
        query = query.order_by(Pilotos.id.desc())
    elif sort == "id_asc":
        query = query.order_by(Pilotos.id.asc())
    else:
        query = query.order_by(Pilotos.nome_piloto.asc())

    # -----------------------------
    # Exportação Excel
    # -----------------------------
    if export == "xlsx":
        rows = query.all()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Pilotos"

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws["A1"] = "Relatório de Pilotos"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(color="6B7280")

        if user_tipo == "uvis":
            ws["A3"] = f"Região (UVIS): {uvis_regiao}"
            ws["A3"].font = Font(color="6B7280")

        start_row = 5 if user_tipo == "uvis" else 4
        headers = ["ID", "Nome", "Região", "Telefone"]

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        for i, p in enumerate(rows, start=start_row + 1):
            values = [p.id, p.nome_piloto, p.regiao or "", format_phone_br(p.telefone or "") or ""]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=i, column=col_idx, value=v)
                cell.border = border
                cell.alignment = center_align if col_idx == 1 else text_align
                if col_idx == 4:
                    cell.number_format = "@"

        last_row = start_row + len(rows)
        last_col = len(headers)

        ws.freeze_panes = ws[f"A{start_row+1}"]
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(last_col)}{max(last_row, start_row)}"

        max_widths = {1: 8, 2: 30, 3: 14, 4: 20}
        for col_idx in range(1, last_col + 1):
            max_len = len(headers[col_idx - 1])
            for r in range(start_row + 1, last_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_widths.get(col_idx, 35))

        zebra_fill = PatternFill("solid", fgColor="F9FAFB")
        for r in range(start_row + 1, last_row + 1):
            if (r - (start_row + 1)) % 2 == 1:
                for ccol in range(1, last_col + 1):
                    ws.cell(row=r, column=ccol).fill = zebra_fill

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"pilotos_{stamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -----------------------------
    # Paginação
    # -----------------------------
    total = query.count()
    total_pages = max(1, math.ceil(total / per_page))
    if page > total_pages:
        page = total_pages

    pilotos_db = query.offset((page - 1) * per_page).limit(per_page).all()

    pilotos = [
        {
            "id": p.id,
            "nome_piloto": p.nome_piloto,
            "regiao": p.regiao or "-",
            "telefone_fmt": format_phone_br(p.telefone or "") or "-",
            "telefone_digits": only_digits(p.telefone or ""),
        }
        for p in pilotos_db
    ]

    filters = {
        "q": q,
        "regiao": regiao,
        "telefone": telefone,
        "sort": sort,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
    }

    return render_template(
        "listar_pilotos.html",
        pilotos=pilotos,
        filters=filters,
        is_admin=(user_tipo == "admin"),
        uvis_regiao=(uvis_regiao if user_tipo == "uvis" else None),
    )


# -------------------------------------------------------------
# EDITAR PILOTO (admin) + atualizar usuário do piloto (login/senha)
# -------------------------------------------------------------
@bp.route("/pilotos/<int:piloto_id>/editar", methods=["GET", "POST"], endpoint="editar_piloto")
@login_required
def editar_piloto(piloto_id):
    # Segurança: só admin
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    piloto = Pilotos.query.get_or_404(piloto_id)

    # usuário de login do piloto (se existir)
    usuario_piloto = Usuario.query.filter_by(piloto_id=piloto.id, tipo_usuario="piloto").first()

    errors = {}
    form = {}

    if request.method == "POST":
        # -----------------------------
        # Campos do piloto
        # -----------------------------
        nome_piloto = (request.form.get("nome_piloto") or "").strip()
        regiao = (request.form.get("regiao") or "").strip().upper()
        telefone = (request.form.get("telefone") or "").strip()
        tel_digits = only_digits(telefone)

        # -----------------------------
        # Campos de acesso (login)
        # -----------------------------
        login = (request.form.get("login") or "").strip()
        senha = (request.form.get("senha") or "").strip()
        senha2 = (request.form.get("senha2") or "").strip()

        # Mantém valores para re-render
        form = {
            "nome_piloto": nome_piloto,
            "regiao": regiao,
            "telefone": telefone,
            "login": login,
            # nunca re-render senha por segurança
        }

        # -----------------------------
        # Validações piloto
        # -----------------------------
        if not nome_piloto:
            errors["nome_piloto"] = "Informe o nome do piloto."

        if regiao and regiao not in REGIOES:
            errors["regiao"] = "Selecione uma região válida."

        if telefone and len(tel_digits) not in (10, 11):
            errors["telefone"] = "Telefone deve ter 10 ou 11 dígitos (com DDD)."

        # duplicidade ignorando o próprio
        if nome_piloto:
            q = Pilotos.query.filter(db.func.lower(Pilotos.nome_piloto) == nome_piloto.lower())
            if tel_digits:
                q = q.filter(Pilotos.telefone == tel_digits)
            q = q.filter(Pilotos.id != piloto.id)
            if q.first():
                errors["nome_piloto"] = "Já existe um piloto com esse nome (e telefone)."

        # -----------------------------
        # Validações acesso
        # -----------------------------
        if not login:
            errors["login"] = "Informe o login do piloto."

        # login único (tirando o próprio usuario_piloto)
        if login:
            q_login = Usuario.query.filter(db.func.lower(Usuario.login) == login.lower())
            if usuario_piloto:
                q_login = q_login.filter(Usuario.id != usuario_piloto.id)
            if q_login.first():
                errors["login"] = "Este login já está em uso. Escolha outro."

        # senha: opcional no editar
        if senha or senha2:
            if len(senha) < 4:
                errors["senha"] = "A senha deve ter pelo menos 4 caracteres."
            if senha != senha2:
                errors["senha2"] = "As senhas não conferem."

        if errors:
            flash("Corrija os campos destacados.", "warning")
            return render_template(
                "editar_piloto.html",
                piloto=piloto,
                form=form,
                errors=errors,
                usuario_piloto=usuario_piloto
            )

        # -----------------------------
        # Salva piloto
        # -----------------------------
        piloto.nome_piloto = nome_piloto
        piloto.regiao = regiao or None
        piloto.telefone = tel_digits or None

        # -----------------------------
        # Salva/Cria usuário do piloto
        # -----------------------------
        if not usuario_piloto:
            # se por algum motivo não existir, cria agora
            usuario_piloto = Usuario(
                nome_uvis=nome_piloto,
                regiao=regiao or None,
                codigo_setor=None,
                login=login,
                tipo_usuario="piloto",
                piloto_id=piloto.id
            )
            # se admin não informou senha ao "editar", força criar uma
            if not senha:
                errors["senha"] = "Defina uma senha para criar o acesso do piloto."
                flash("Corrija os campos destacados.", "warning")
                return render_template(
                    "editar_piloto.html",
                    piloto=piloto,
                    form=form,
                    errors=errors,
                    usuario_piloto=usuario_piloto
                )
            usuario_piloto.set_senha(senha)
            db.session.add(usuario_piloto)
        else:
            # atualiza dados básicos do usuario
            usuario_piloto.nome_uvis = nome_piloto
            usuario_piloto.regiao = regiao or None
            usuario_piloto.login = login

            # troca senha somente se veio preenchida
            if senha:
                usuario_piloto.set_senha(senha)

        db.session.commit()

        flash("Piloto atualizado com sucesso!", "success")
        return redirect(url_for("main.listar_pilotos"))

    # -----------------------------
    # GET: valores default
    # -----------------------------
    form = {
        "nome_piloto": piloto.nome_piloto,
        "regiao": (piloto.regiao or ""),
        "telefone": format_phone_br(piloto.telefone or ""),
        "login": (usuario_piloto.login if usuario_piloto else ""),
    }

    return render_template(
        "editar_piloto.html",
        piloto=piloto,
        form=form,
        errors=errors,
        usuario_piloto=usuario_piloto
    )


# -------------------------------------------------------------
# DELETAR PILOTO (admin) + deletar usuário do piloto (se existir)
# -------------------------------------------------------------
@bp.route("/pilotos/<int:piloto_id>/deletar", methods=["POST"], endpoint="deletar_piloto")
@login_required
def deletar_piloto(piloto_id):
    if getattr(current_user, "tipo_usuario", None) != "admin":
        abort(403)

    piloto = Pilotos.query.get_or_404(piloto_id)

    # apaga usuário(s) vinculados a esse piloto
    Usuario.query.filter_by(piloto_id=piloto.id, tipo_usuario="piloto").delete(synchronize_session=False)

    db.session.delete(piloto)
    db.session.commit()

    flash("Piloto excluído com sucesso.", "success")
    return redirect(url_for("main.listar_pilotos"))


@bp.route('/piloto/os')
@login_required
@roles_required('piloto')
def piloto_os():
    if not current_user.piloto_id:
        flash("Piloto sem vínculo cadastrado.", "danger")
        return redirect(url_for('main.dashboard'))

    status_ok = ["APROVADA", "APROVADA COM RECOMENDAÇÕES", "APROVADO", "APROVADO COM RECOMENDAÇÕES"]

    query = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter(
            Solicitacao.piloto_id == current_user.piloto_id,
            Solicitacao.status.in_(status_ok)
        )
    )

    filtro_data = request.args.get("data")  # ex: 2026-01
    uvis_id = request.args.get("uvis_id")
    query = aplicar_filtros_base(query, filtro_data, uvis_id)

    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(
        Solicitacao.data_agendamento.asc(),
        Solicitacao.hora_agendamento.asc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        "piloto_os.html",
        pedidos=paginacao.items,
        paginacao=paginacao,
        status_ok=status_ok
    )


@bp.route('/piloto/os/<int:os_id>/concluir', methods=['POST'])
@login_required
@roles_required('piloto')
def piloto_concluir_os(os_id):

    s = Solicitacao.query.get_or_404(os_id)

    if s.piloto_id != current_user.piloto_id:
        flash("Você não pode alterar esta OS.", "danger")
        return redirect(url_for('main.piloto_os'))

    status_ok = ["APROVADO", "APROVADO COM RECOMENDAÇÕES", "APROVADA", "APROVADA COM RECOMENDAÇÕES"]
    if s.status not in status_ok:
        flash("A OS não está aprovada.", "warning")
        return redirect(url_for('main.piloto_os'))

    s.status = "CONCLUÍDO"
    db.session.commit()

    flash("Ordem de serviço concluída com sucesso.", "success")
    return redirect(url_for('main.piloto_os'))


